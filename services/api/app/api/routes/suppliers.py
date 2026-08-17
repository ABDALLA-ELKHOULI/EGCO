# -*- coding: utf-8 -*-
from typing import List, Optional
import datetime as dt
import io
from decimal import Decimal
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.deps import parse_date
from app.db import models
from app.db.session import get_session
from app.domain.payables import D, DELAY_BUCKETS, parse_term, money
from app.schemas.common import SupplierIn, SupplierUpdate
from app.services import party_projects as PP
from app.services import payables_service as PS
from app.utils.arabic import contains_ar

router = APIRouter()


def _activity(db: Session, supplier_id: str):
    dates = []
    for i in db.query(models.Invoice.date).filter_by(supplier_id=supplier_id,
                                                       deleted_at=None).all():
        dates.append(i[0])
    for p in db.query(models.Payment.date).filter_by(supplier_id=supplier_id,
                                                       deleted_at=None).all():
        dates.append(p[0])
    if not dates:
        return None, None
    return min(dates).isoformat(), max(dates).isoformat()


_VALID_STATUSES = ('awaiting_date', 'overdue', 'due_soon', 'open', 'clear')

_VALID_DELAY = tuple(k for k, _, _ in DELAY_BUCKETS) + ('none',)

#: أعمدة الترتيب — المفتاح كما يرسله الجدول، والدالة التي تستخرج قيمة الفرز.
#: Sorting on the server (over the full filtered set) is what keeps the totals row
#: honest: sort a page in the browser and the totals silently describe a different set.
_SORT_KEYS = {
    'name': lambda r: r['name'] or '',
    'account': lambda r: r['account'] or '',
    'project': lambda r: r['project'] or '',
    'term': lambda r: r.get('termDays') if r.get('termKind') == 'days' else -1,
    'status': lambda r: _VALID_STATUSES.index(r['status']) if r['status'] in _VALID_STATUSES else 99,
    'outstanding': lambda r: r['outstanding'],
    'overdue': lambda r: r['overdue'],
    'delay': lambda r: r['delay']['days'],
    'delayAmount': lambda r: r['delay']['amount'],
    'lastPaymentDate': lambda r: (r['lastPayment'] or {}).get('date') or '',
    'lastPaymentAmount': lambda r: (r['lastPayment'] or {}).get('amount') or 0,
}


def _parse_filter_date(s: Optional[str], label: str) -> Optional[dt.date]:
    if not s:
        return None
    try:
        return dt.date.fromisoformat(s)
    except ValueError:
        raise HTTPException(422, detail=f'{label} غير صالح: {s}')


@router.get('')
def list_suppliers(q: Optional[str] = Query(None),
                   project: Optional[str] = Query(None),
                   status: Optional[str] = Query(None),
                   date_from: Optional[str] = Query(None),
                   date_to: Optional[str] = Query(None),
                   min_outstanding: Optional[float] = Query(None),
                   max_outstanding: Optional[float] = Query(None),
                   overdue_only: bool = Query(False),
                   has_data: Optional[bool] = Query(None),
                   delay: Optional[str] = Query(None),
                   min_delay_days: Optional[int] = Query(None),
                   max_delay_days: Optional[int] = Query(None),
                   sort: Optional[str] = Query(None),
                   dir: str = Query('asc'),
                   db: Session = Depends(get_session)) -> dict:
    """قائمة الموردين مع حالتهم — filtering happens after positions are computed,
    because status depends on the calculation, not on a stored column.

    Totals always reflect the applied filters, over the FULL filtered set (not just the
    rows shown) — one code path computes both."""
    if status is not None and status not in _VALID_STATUSES:
        raise HTTPException(422, detail=f'قيمة حالة غير صالحة: {status} — '
                                        f'المسموح: {", ".join(_VALID_STATUSES)}')
    if delay is not None and delay not in _VALID_DELAY:
        raise HTTPException(422, detail=f'شريحة تأخر غير صالحة: {delay} — '
                                        f'المسموح: {", ".join(_VALID_DELAY)}')
    if sort is not None and sort not in _SORT_KEYS:
        raise HTTPException(422, detail=f'عمود ترتيب غير صالح: {sort} — '
                                        f'المسموح: {", ".join(_SORT_KEYS)}')
    if dir not in ('asc', 'desc'):
        raise HTTPException(422, detail=f'اتجاه ترتيب غير صالح: {dir}')
    df = _parse_filter_date(date_from, 'تاريخ البداية')
    dtt = _parse_filter_date(date_to, 'تاريخ النهاية')
    if df is not None and dtt is not None and df > dtt:
        raise HTTPException(422, detail='تاريخ البداية يجب أن يسبق تاريخ النهاية')
    if (min_outstanding is not None and max_outstanding is not None
            and min_outstanding > max_outstanding):
        raise HTTPException(422, detail='الحد الأدنى للرصيد يجب ألا يتجاوز الحد الأقصى')

    ps = PS.positions(db, include_empty=True)
    rows = []
    matched = []
    supplier_ids = {row.account: row.id for row in
                    db.query(models.Supplier).filter(models.Supplier.deleted_at.is_(None)).all()}
    # التصفية بمشروع تعني «ينتمي إليه ضمن لائحته» لا «يساويه» — مورد على ثلاثة
    # مشاريع يجب أن يظهر تحت الثلاثة، لا تحت مشروعه الأساسي (project) فقط.
    ids_in_project = PP.parties_in_project(db, PP.SUPPLIER, project) if project else None
    for p in ps:
        st = PS.status_of(p)
        if status and st != status:
            continue
        if ids_in_project is not None and supplier_ids.get(p.supplier.account) not in ids_in_project:
            continue
        if q:
            # المقارنة مُطبَّعة عربياً (الهمزات/التاء المربوطة/الياء الفارسية...) —
            # اسم المورد كما كتبه نظام الحسابات القديم قد يختلف إملائياً حرفاً واحداً
            # عمّا يكتبه المستخدم بحثاً وهو يقصد نفس المورد بالضبط. انظر app/utils/arabic.py.
            if not contains_ar(p.supplier.name, q) and not contains_ar(p.supplier.account, q):
                continue
        if min_outstanding is not None and p.outstanding < Decimal(str(min_outstanding)):
            continue
        if max_outstanding is not None and p.outstanding > Decimal(str(max_outstanding)):
            continue
        if overdue_only and p.overdue <= 0:
            continue
        # التصفية بالشريحة تعني «له مبلغ متأخر فيها» لا «أسوأ تأخره فيها» — سؤالك
        # «أرني المتأخر ٣١–٦٠ يوماً» يقصد المال الواقع هناك، وقد يملكه مورد أسوأ
        # تأخره ستة أشهر. لذلك يظهر المورد في كل شريحة له فيها مال.
        if delay is not None:
            if delay == 'none':
                if p.delay.days > 0:
                    continue
            elif p.delay.by_bucket.get(delay, Decimal('0')) <= 0:
                continue
        if min_delay_days is not None and p.delay.days < min_delay_days:
            continue
        if max_delay_days is not None and p.delay.days > max_delay_days:
            continue
        if has_data is not None:
            has_movement = bool(p.invoices or p.payments)
            if has_movement != has_data:
                continue
        sid = supplier_ids.get(p.supplier.account)
        d = PS.position_json(p, projects=PP.projects_of(db, PP.SUPPLIER, sid) if sid else [])
        d['status'] = st
        first_act, last_act = _activity(db, sid) if sid else (None, None)
        d['firstActivity'] = first_act
        d['lastActivity'] = last_act
        rows.append(d)
        matched.append(p)

    if sort:
        # الاسم فاصل التعادل دائماً — بدونه يتبدّل ترتيب المتساويات بين طلب وآخر
        # فيبدو الجدول وكأنه يتحرك بلا سبب.
        rows.sort(key=lambda r: r['name'])
        rows.sort(key=_SORT_KEYS[sort], reverse=(dir == 'desc'))
    else:
        rows.sort(key=lambda r: (-r['overdue'], -r['outstanding'], r['name']))
    # كل المشاريع المعروفة من party_projects — يشمل مشاريع إضافية لمورد له أكثر من
    # مشروع، لا العمود المفرد فقط.
    projects = PP.all_projects(db)
    # Sum the Decimal positions, not the already-rounded per-row floats — summing
    # money()-rounded floats can drift by fractions of a piaster from the exact total
    # (e.g. 5611014.100000001), disagreeing with dashboard/overview/projects which all
    # sum Decimals before rounding once at the boundary.
    zero = Decimal('0')
    totals = dict(
        count=len(matched),
        invoiced=money(sum((p.total_invoiced for p in matched), zero)),
        paid=money(sum((p.total_paid for p in matched), zero)),
        outstanding=money(sum((p.outstanding for p in matched), zero)),
        overdue=money(sum((p.overdue for p in matched), zero)),
        dueWithin7=money(sum((p.due_within_7 for p in matched), zero)),
        creditBalances=money(sum((p.credit_balance for p in matched), zero)),
        # توزيع المتأخر على الشرائح — للمجموعة المصفّاة كاملة لا للصفحة المعروضة.
        delayByBucket={k: money(sum((p.delay.by_bucket.get(k, zero) for p in matched), zero))
                       for k, _, _ in DELAY_BUCKETS},
        delayed=money(sum((p.delay.amount for p in matched), zero)),
    )
    if df is not None or dtt is not None:
        opening = invoiced_p = paid_p = closing = zero
        for p in matched:
            b = PS.period_breakdown(p, df, dtt)
            opening += b['opening']
            invoiced_p += b['invoiced_in_period']
            paid_p += b['paid_in_period']
            closing += b['closing']
        totals['openingBalance'] = money(opening)
        totals['invoicedInPeriod'] = money(invoiced_p)
        totals['paidInPeriod'] = money(paid_p)
        totals['closingBalance'] = money(closing)

    filters_applied = dict(q=q, project=project, status=status, dateFrom=date_from,
                           dateTo=date_to, minOutstanding=min_outstanding,
                           maxOutstanding=max_outstanding, overdueOnly=overdue_only,
                           hasData=has_data)
    return dict(count=len(rows), rows=rows, projects=projects, totals=totals,
               filtersApplied=filters_applied)


@router.get('/export.xlsx')
def export_suppliers_xlsx(q: Optional[str] = Query(None),
                          project: Optional[str] = Query(None),
                          status: Optional[str] = Query(None),
                          date_from: Optional[str] = Query(None),
                          date_to: Optional[str] = Query(None),
                          min_outstanding: Optional[float] = Query(None),
                          max_outstanding: Optional[float] = Query(None),
                          overdue_only: bool = Query(False),
                          has_data: Optional[bool] = Query(None),
                          delay: Optional[str] = Query(None),
                          min_delay_days: Optional[int] = Query(None),
                          max_delay_days: Optional[int] = Query(None),
                          sort: Optional[str] = Query(None),
                          dir: str = Query('asc'),
                          db: Session = Depends(get_session)):
    """تصدير لائحة الموردين — بنفس التصفية المطبَّقة على الشاشة بالضبط، لا الدفتر
    كاملاً. تستدعي list_suppliers مباشرةً (كما تفعل reports.py مع project_summary)
    حتى يبقى مسار تصفية واحد يُثق بنتائجه في الشاشة والملف معاً.
    """
    data = list_suppliers(q=q, project=project, status=status, date_from=date_from,
                          date_to=date_to, min_outstanding=min_outstanding,
                          max_outstanding=max_outstanding, overdue_only=overdue_only,
                          has_data=has_data, delay=delay, min_delay_days=min_delay_days,
                          max_delay_days=max_delay_days, sort=sort, dir=dir, db=db)

    wb = Workbook()
    ws = wb.active
    ws.title = 'الموردون'
    ws.sheet_view.rightToLeft = True
    ws.append(['رقم الحساب', 'الاسم', 'المشروع', 'الحالة', 'المديونية المفتوحة',
              'المتأخر', 'آخر دفعة — التاريخ', 'آخر دفعة — المبلغ'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in data['rows']:
        lp = r.get('lastPayment') or {}
        delay_amt = (r.get('delay') or {}).get('amount', 0) if r.get('delay') else 0
        ws.append([r['account'], r['name'], r.get('project', ''), r['status'],
                  r['outstanding'], delay_amt, lp.get('date', ''), lp.get('amount', '')])
    t = data['totals']
    ws.append(['الإجمالي', '', '', '', t['outstanding'], t['delayed'], '', ''])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for row in range(2, ws.max_row + 1):
        for col in (5, 6, 8):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    today = dt.date.today()
    ascii_name = f'EGCO-suppliers-{today:%Y%m%d}.xlsx'
    encoded = quote(f'EGCO-الموردون-{today:%Y%m%d}.xlsx', safe='')
    headers = {'Content-Disposition':
              f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded}"}
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )


class SmartAllocationSettingIn(BaseModel):
    enabled: bool


class AllocationLineIn(BaseModel):
    """سطر تخصيص واحد — فاتورة ومبلغ، أو invoiceId=None لجزء «على الحساب»."""
    invoiceId: Optional[str] = None
    amount: float


class PaymentAllocationIn(BaseModel):
    lines: List[AllocationLineIn]


# ---------------------------------------------------------------- تخصيص الدفعات (opt-in)
# مسارات ثابتة يجب أن تُسجَّل قبل `/{account}` كي لا يبتلعها كمسار حساب.

@router.get('/settings/payment-allocation')
def get_payment_allocation_setting(db: Session = Depends(get_session)) -> dict:
    """هل ميزة «تخصيص الدفعات» مفعّلة؟ افتراضياً متوقفة — انظر شرح كامل الميزة
    في app/db/models.py (PaymentAllocation) وapp/domain/payables.py (allocate_smart)."""
    return dict(enabled=PS.is_smart_allocation_enabled(db),
               pendingCount=PS.count_pending_allocations(db))


@router.put('/settings/payment-allocation')
def set_payment_allocation_setting(body: SmartAllocationSettingIn,
                                   db: Session = Depends(get_session)) -> dict:
    """تشغيل/إيقاف الميزة — تغيير هذا الإعداد لا يمحو أي قرار محفوظ سابقاً؛
    إيقافها يعيد الحساب فوراً إلى allocate_fifo القديم بلا أي أثر للجدول الجديد."""
    PS.set_smart_allocation_enabled(db, body.enabled)
    return dict(enabled=body.enabled, pendingCount=PS.count_pending_allocations(db))


@router.get('/payment-allocation/pending-count')
def payment_allocation_pending_count(db: Session = Depends(get_session)) -> dict:
    """عدد الدفعات بانتظار التخصيص عبر كل الموردين — لعرض «٣ دفعات بانتظار
    التخصيص» في مكان واحد بدل تذكّره من كل شاشة مورد على حدة."""
    return dict(count=PS.count_pending_allocations(db))


@router.post('/{account}/payments/{payment_id}/allocate')
def allocate_payment(account: str, payment_id: str, body: PaymentAllocationIn,
                     db: Session = Depends(get_session)) -> dict:
    """يخصص دفعة (معلَّقة أو مُخصَّصة سابقاً) لفاتورة/فواتير أو «على الحساب» —
    قرار المستخدم دائماً، يُخزَّن ويحل محل أي قرار سابق لنفس الدفعة، ويمكن
    تعديله لاحقاً بنداء آخر لنفس المسار (استبدال كامل، لا دمج)."""
    if not PS.is_smart_allocation_enabled(db):
        raise HTTPException(409, detail='ميزة تخصيص الدفعات متوقفة — فعّلها أولاً من الإعدادات')
    supplier = db.query(models.Supplier).filter_by(account=account).filter(
        models.Supplier.deleted_at.is_(None)).one_or_none()
    if supplier is None:
        raise HTTPException(404, detail=f'لا يوجد مورد بالحساب {account}')
    payment = db.query(models.Payment).filter_by(
        id=payment_id, supplier_id=supplier.id, deleted_at=None).one_or_none()
    if payment is None:
        raise HTTPException(404, detail='لا توجد دفعة بهذا المعرف لهذا المورد')
    if not body.lines:
        raise HTTPException(422, detail='يجب تحديد سطر تخصيص واحد على الأقل — استخدم '
                                        'invoiceId فارغاً لتخصيصها بالكامل «على الحساب»')

    # المجموع يجب أن يساوي مبلغ الدفعة تماماً (بسماحية القروش) — تخصيص جزئي بلا
    # تفسير لبقية المبلغ يُعيد نفس مشكلة «الفرق يُعتبر سداداً» التي بدأت منها الميزة.
    total = sum((D(l.amount) for l in body.lines), D('0'))
    if abs(total - D(payment.amount)) > D('0.01'):
        raise HTTPException(422, detail=f'مجموع أسطر التخصيص ({money(total)}) لا يساوي '
                                        f'مبلغ الدفعة ({money(payment.amount)})')
    for l in body.lines:
        if l.amount <= 0:
            raise HTTPException(422, detail='مبلغ كل سطر يجب أن يكون أكبر من صفر')
        if l.invoiceId is not None:
            inv = db.query(models.Invoice).filter_by(
                id=l.invoiceId, supplier_id=supplier.id, deleted_at=None).one_or_none()
            if inv is None:
                raise HTTPException(422, detail=f'الفاتورة {l.invoiceId} لا تتبع هذا المورد')

    PS.save_payment_allocation(
        db, payment_id, [dict(invoiceId=l.invoiceId, amount=l.amount) for l in body.lines])
    return dict(saved=True)


@router.delete('/{account}/payments/{payment_id}/allocate')
def delete_payment_allocation(account: str, payment_id: str,
                              db: Session = Depends(get_session)) -> dict:
    """يمحو قرار تخصيص سابق — تعود الدفعة تلقائياً لقائمة المراجعة إن كانت
    غامضة، أو تُخصَّص تلقائياً من جديد إن لم تكن (نفس منطق allocate_smart)."""
    supplier = db.query(models.Supplier).filter_by(account=account).filter(
        models.Supplier.deleted_at.is_(None)).one_or_none()
    if supplier is None:
        raise HTTPException(404, detail=f'لا يوجد مورد بالحساب {account}')
    payment = db.query(models.Payment).filter_by(
        id=payment_id, supplier_id=supplier.id, deleted_at=None).one_or_none()
    if payment is None:
        raise HTTPException(404, detail='لا توجد دفعة بهذا المعرف لهذا المورد')
    PS.clear_payment_allocation(db, payment_id)
    return dict(cleared=True)


@router.get('/{account}')
def supplier_detail(account: str,
                    date_from: Optional[str] = Query(None),
                    date_to: Optional[str] = Query(None),
                    db: Session = Depends(get_session)) -> dict:
    """كشف مورد — invoices, payments, ageing."""
    # include_empty=True: مورد بلا حركة حيّة مورد قائم، لا مورد غير موجود. بدونها
    # كان حذف ملفه يجعل صفحته تقول «لا يوجد مورد بالحساب ٢١١٠٨٠٨» — وهو موجود،
    # وكشفه كله ينتظر إعادة الرفع. رسالةٌ تنفي وجود شيء قائم أسوأ من صفحة فارغة.
    ps = PS.positions(db, account=account, include_empty=True)
    if not ps:
        raise HTTPException(404, detail=f'لا يوجد مورد بالحساب {account}')
    df = parse_date(date_from, 'تاريخ البداية')
    dtt = parse_date(date_to, 'تاريخ النهاية')
    sid = db.query(models.Supplier.id).filter_by(account=account).scalar()
    d = PS.position_json(ps[0], detail=True, date_from=df, date_to=dtt,
                         projects=PP.projects_of(db, PP.SUPPLIER, sid) if sid else [])
    d['status'] = PS.status_of(ps[0])
    return d


@router.post('', status_code=201)
def create_supplier(body: SupplierIn, db: Session = Depends(get_session)) -> dict:
    """إضافة مورد يدوياً."""
    exists = db.query(models.Supplier).filter_by(account=body.account).one_or_none()
    if exists is not None and exists.deleted_at is None:
        raise HTTPException(409, detail=f'يوجد مورد بالفعل بالحساب {body.account}')
    term = parse_term(body.term)
    if exists is not None:
        # previously soft-deleted — resurrect
        row = exists
        row.deleted_at = None
    else:
        row = models.Supplier(account=body.account)
        db.add(row)
    row.name = body.name
    row.project = body.project
    row.term_raw = term.raw
    row.term_kind = term.kind
    row.term_days = term.days
    db.flush()  # نحتاج row.id قبل set_projects
    # body.projects لو أُرسل يحل محل project المفرد كأساس؛ لو لم يُرسل نبذر
    # اللائحة من project حتى لا يبقى المورد الجديد بلا مشاريع في party_projects.
    projects_in = body.projects if body.projects is not None else (
        [body.project] if body.project else [])
    projects = PP.set_projects(db, PP.SUPPLIER, row.id, projects_in)
    row.project = PP.primary(projects)
    db.commit()
    db.refresh(row)
    return dict(account=row.account, name=row.name, project=row.project,
                term=row.term_raw or 'كاش', termKind=row.term_kind, termDays=row.term_days,
                projects=projects)


@router.put('/{account}')
def update_supplier(account: str, body: SupplierUpdate,
                    db: Session = Depends(get_session)) -> dict:
    """تعديل بيانات مورد — رقم الحساب ثابت."""
    row = db.query(models.Supplier).filter_by(account=account).filter(
        models.Supplier.deleted_at.is_(None)).one_or_none()
    if row is None:
        raise HTTPException(404, detail=f'لا يوجد مورد بالحساب {account}')
    if body.name is not None:
        row.name = body.name
    if body.term is not None:
        term = parse_term(body.term)
        row.term_raw = term.raw
        row.term_kind = term.kind
        row.term_days = term.days
    # body.projects=None يعني «لا تغيّر» — تعديل جزئي (اسم أو مدة فقط) لا يمحو
    # مشاريع المورد بصمت (انظر party_projects.set_projects). لو أُرسل project المفرد
    # وحده بلا projects، نعامله كطلب استبدال باللائحة [project] — سلوك متوافق مع
    # الشاشات القديمة التي لا تعرف بعد بالمشاريع المتعددة.
    projects_in = body.projects
    if projects_in is None and body.project is not None:
        projects_in = [body.project] if body.project else []
    projects = PP.set_projects(db, PP.SUPPLIER, row.id, projects_in)
    if projects_in is not None:
        row.project = PP.primary(projects)
    db.commit()
    db.refresh(row)
    return dict(account=row.account, name=row.name, project=row.project,
                term=row.term_raw or 'كاش', termKind=row.term_kind, termDays=row.term_days,
                projects=projects)


@router.delete('/{account}')
def delete_supplier(account: str, force: bool = Query(False),
                    db: Session = Depends(get_session)) -> dict:
    """حذف مورد (حذف منطقي). إن كان له فواتير/دفعات، يُرفض إلا مع force=true."""
    row = db.query(models.Supplier).filter_by(account=account).filter(
        models.Supplier.deleted_at.is_(None)).one_or_none()
    if row is None:
        raise HTTPException(404, detail=f'لا يوجد مورد بالحساب {account}')

    has_invoices = db.query(models.Invoice).filter_by(
        supplier_id=row.id, deleted_at=None).first() is not None
    has_payments = db.query(models.Payment).filter_by(
        supplier_id=row.id, deleted_at=None).first() is not None

    if (has_invoices or has_payments) and not force:
        raise HTTPException(409, detail='لا يمكن حذف المورد لوجود فواتير أو دفعات مسجلة له — '
                                        'استخدم force=true للحذف مع الإبقاء على السجلات')

    row.deleted_at = dt.datetime.now(dt.timezone.utc)
    db.commit()
    return dict(deleted=True)
