# -*- coding: utf-8 -*-
"""مسارات المقاولين — list / detail / CRUD for entries, claims, guarantees.

Conventions follow suppliers.py: soft delete everywhere, force=true to delete a
contractor that still has records, Arabic HTTP details, camelCase JSON out.
"""
import datetime as dt
import io
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.db import models
from app.db.session import get_session
from app.domain import contractors as C
from app.schemas.contractors import (ClaimIn, ClaimUpdate, ContractorIn,
                                     ContractorUpdate, EntryIn, EntryUpdate,
                                     GuaranteeIn, GuaranteeUpdate)
from app.services import contractors_service as CS
from app.services import party_projects as PP

router = APIRouter()


def _get_contractor(db: Session, code: str) -> models.Contractor:
    row = db.query(models.Contractor).filter_by(code=code).filter(
        models.Contractor.deleted_at.is_(None)).one_or_none()
    if row is None:
        raise HTTPException(404, detail=f'لا يوجد مقاول بالكود {code}')
    return row


def _date(s: str) -> dt.date:
    try:
        return dt.date.fromisoformat(s)
    except ValueError:
        raise HTTPException(422, detail=f'تاريخ غير صالح: {s}')


def _commit_or_409(db: Session, detail: str) -> None:
    """commit that maps a unique-constraint violation to a 409 instead of a 500."""
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, detail=detail)


# ---------------------------------------------------------------- list / detail

_VALID_DIRECTIONS = ('owed_to_them', 'owed_to_us', 'balanced')

# لا توجد شريحة «تأخر» للمقاولين: حركات دفتر المقاول قيود مدين/دائن (مستخلصات
# ودفعات) بلا تاريخ استحقاق يُقاس منه التأخر — بخلاف فواتير الموردين التي لها
# مدة سداد وتاريخ استحقاق مشتق. compute_delay في app/domain/payables.py يعمل على
# Invoice.due_date، وهذا الحقل غير موجود أصلاً في app/domain/contractors.py.
# افتراع تاريخ استحقاق من لا شيء كان يُخرج رقماً كاذباً، فتُركت التصفية هنا مقتصرة
# على ما تدعمه البيانات فعلاً: البحث والمشروع والاتجاه والضمانات، بالإضافة للترتيب.
CS_VALID_SORT = tuple(CS.CONTRACTOR_SORT_KEYS.keys())


@router.get('')
def list_contractors(q: Optional[str] = Query(None),
                     project: Optional[str] = Query(None),
                     direction: Optional[str] = Query(None),
                     has_guarantees: Optional[bool] = Query(None),
                     sort: Optional[str] = Query(None),
                     dir: str = Query('asc'),
                     db: Session = Depends(get_session)) -> dict:
    """قائمة المقاولين — sorted most-negative balance first (أكبر مستحقات لهم أولاً)
    ما لم يُطلب ترتيب صريح عبر sort/dir."""
    if direction is not None and direction not in _VALID_DIRECTIONS:
        raise HTTPException(422, detail=f'قيمة اتجاه غير صالحة: {direction} — '
                                        f'المسموح: {", ".join(_VALID_DIRECTIONS)}')
    if sort is not None and sort not in CS_VALID_SORT:
        raise HTTPException(422, detail=f'عمود ترتيب غير صالح: {sort} — '
                                        f'المسموح: {", ".join(CS_VALID_SORT)}')
    if dir not in ('asc', 'desc'):
        raise HTTPException(422, detail=f'اتجاه ترتيب غير صالح: {dir}')
    return CS.contractors_list_json(db, q=q, project=project, direction=direction,
                                    has_guarantees=has_guarantees, sort=sort, dir=dir)


@router.get('/overview')
def contractors_overview(db: Session = Depends(get_session)) -> dict:
    """لوحة نظرة المقاولين — التوزيع على المشاريع، أكبر المستحقات، الضمانات
    المستقلة (216)، واختلافات الرصيد مقابل أحدث تقرير مديونيات مجمّع تم استيراده.
    مسار مستقل عن `list_contractors` أعلاه: هذا مُجمَّع على مستوى الشركة كاملة
    ولا يتأثر بأي تصفية شاشة القائمة (بحث/مشروع/اتجاه)."""
    return CS.contractors_overview_json(db)


@router.get('/export.xlsx')
def export_contractors_xlsx(q: Optional[str] = Query(None),
                            project: Optional[str] = Query(None),
                            direction: Optional[str] = Query(None),
                            has_guarantees: Optional[bool] = Query(None),
                            sort: Optional[str] = Query(None),
                            dir: str = Query('asc'),
                            db: Session = Depends(get_session)):
    """تصدير لائحة المقاولين — بنفس التصفية المطبَّقة على الشاشة، لا الدفتر كاملاً.
    نفس نمط export_suppliers_xlsx: يستدعي list_contractors مباشرةً فيبقى مساراً
    واحداً للتصفية يصف الشاشة والملف معاً."""
    data = list_contractors(q=q, project=project, direction=direction,
                            has_guarantees=has_guarantees, sort=sort, dir=dir, db=db)

    wb = Workbook()
    ws = wb.active
    ws.title = 'المقاولون'
    ws.sheet_view.rightToLeft = True
    ws.append(['كود المقاول', 'الاسم', 'المشاريع', 'المحمّل عليه', 'المدفوع',
              'خصومات وتحميلات', 'الرصيد', 'ضمانات محتجزة'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in data['rows']:
        ws.append([r.get('code', ''), r.get('name', ''),
                  '، '.join(r.get('projects') or []),
                  r.get('duesTotal', 0), r.get('paidTotal', 0),
                  r.get('deductionsTotal', 0), r.get('balance', 0),
                  r.get('retentionHeld', 0)])
    t = data['totals']
    ws.append(['الإجمالي', '', '', t.get('claimsTotal', 0), t.get('paidTotal', 0),
              t.get('deductionsTotal', 0), t.get('balance', 0), t.get('retentionHeld', 0)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for row in range(2, ws.max_row + 1):
        for col in (4, 5, 6, 7, 8):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    today = dt.date.today()
    ascii_name = f'EGCO-contractors-{today:%Y%m%d}.xlsx'
    encoded = quote(f'EGCO-المقاولون-{today:%Y%m%d}.xlsx', safe='')
    headers = {'Content-Disposition':
              f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded}"}
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )


@router.get('/{code}')
def contractor_detail(code: str, db: Session = Depends(get_session)) -> dict:
    """كشف مقاول كامل — الحركات والمستخلصات والضمانات وتوزيع المشاريع."""
    row = _get_contractor(db, code)
    return CS.contractor_detail_json(row, projects=PP.projects_of(db, PP.CONTRACTOR, row.id))


# ---------------------------------------------------------------- contractor CRUD

@router.post('', status_code=201)
def create_contractor(body: ContractorIn, db: Session = Depends(get_session)) -> dict:
    """إضافة مقاول يدوياً — resurrects a soft-deleted code, like suppliers."""
    exists = db.query(models.Contractor).filter_by(code=body.code).one_or_none()
    if exists is not None and exists.deleted_at is None:
        raise HTTPException(409, detail=f'يوجد مقاول بالفعل بالكود {body.code}')
    if exists is not None:
        row = exists
        row.deleted_at = None
    else:
        row = models.Contractor(code=body.code)
        db.add(row)
    row.name = body.name
    row.phone = body.phone
    row.notes = body.notes
    row.default_retention_rate = body.defaultRetentionRate
    row.default_guarantee_days = body.defaultGuaranteeDays
    db.flush()  # نحتاج row.id قبل set_projects
    projects = PP.set_projects(db, PP.CONTRACTOR, row.id, body.projects)
    db.commit()
    db.refresh(row)
    return CS.contractor_row_json(row, projects=projects)


@router.put('/{code}')
def update_contractor(code: str, body: ContractorUpdate,
                      db: Session = Depends(get_session)) -> dict:
    """تعديل بيانات مقاول — الكود ثابت."""
    row = _get_contractor(db, code)
    if body.name is not None:
        row.name = body.name
    if body.phone is not None:
        row.phone = body.phone
    if body.notes is not None:
        row.notes = body.notes
    if body.defaultRetentionRate is not None:
        row.default_retention_rate = body.defaultRetentionRate
    if body.defaultGuaranteeDays is not None:
        row.default_guarantee_days = body.defaultGuaranteeDays
    # body.projects=None يعني «لا تغيّر» — تعديل جزئي (اسم فقط) لا يمحو المشاريع.
    projects = PP.set_projects(db, PP.CONTRACTOR, row.id, body.projects)
    db.commit()
    db.refresh(row)
    return CS.contractor_row_json(row, projects=projects)


@router.delete('/{code}')
def delete_contractor(code: str, force: bool = Query(False),
                      db: Session = Depends(get_session)) -> dict:
    """حذف مقاول (حذف منطقي). إن كانت له حركات مسجلة يُرفض إلا مع force=true."""
    row = _get_contractor(db, code)
    has_entries = db.query(models.ContractorEntry).filter_by(
        contractor_id=row.id, deleted_at=None).first() is not None
    if has_entries and not force:
        raise HTTPException(409, detail='لا يمكن حذف المقاول لوجود حركات مسجلة له — '
                                        'استخدم force=true للحذف مع الإبقاء على السجلات')
    row.deleted_at = dt.datetime.now(dt.timezone.utc)
    db.commit()
    return dict(deleted=True)


# ---------------------------------------------------------------- entries

def _validate_sides(debit: float, credit: float) -> None:
    # Exactly one side of a ledger line may carry a value — that is what keeps
    # balance = Σdebit − Σcredit meaningful.
    if not ((debit > 0) ^ (credit > 0)) or debit < 0 or credit < 0:
        raise HTTPException(422, detail='يجب إدخال قيمة موجبة في مدين أو دائن (واحد فقط)')


@router.post('/{code}/entries', status_code=201)
def create_entry(code: str, body: EntryIn, db: Session = Depends(get_session)) -> dict:
    """حركة يدوية على دفتر المقاول."""
    row = _get_contractor(db, code)
    _validate_sides(body.debit, body.credit)
    kind = body.kind or C.classify_entry(body.description)
    entry = models.ContractorEntry(
        contractor_id=row.id, date=_date(body.date), debit=body.debit,
        credit=body.credit, description=body.description, kind=kind,
        claim_no=body.claimNo if body.claimNo is not None
                 else C.extract_claim_no(body.description),
        project=body.project if body.project is not None
                else C.detect_project(body.description, CS.known_projects(db)),
        source='manual')
    db.add(entry)
    _commit_or_409(db, 'توجد حركة مطابقة تماماً (نفس التاريخ والمبلغ والوصف) — '
                       'عدّل الوصف أو المبلغ للتمييز بينهما')
    db.refresh(entry)
    return CS.entry_json(entry)


def _get_entry(db: Session, row: models.Contractor, entry_id: str) -> models.ContractorEntry:
    e = db.query(models.ContractorEntry).filter_by(
        id=entry_id, contractor_id=row.id).filter(
        models.ContractorEntry.deleted_at.is_(None)).one_or_none()
    if e is None:
        raise HTTPException(404, detail='لا توجد حركة بهذا المعرف')
    return e


@router.put('/{code}/entries/{entry_id}')
def update_entry(code: str, entry_id: str, body: EntryUpdate,
                 db: Session = Depends(get_session)) -> dict:
    row = _get_contractor(db, code)
    e = _get_entry(db, row, entry_id)
    if body.date is not None:
        e.date = _date(body.date)
    if body.debit is not None:
        e.debit = body.debit
    if body.credit is not None:
        e.credit = body.credit
    _validate_sides(e.debit or 0, e.credit or 0)
    if body.description is not None:
        e.description = body.description
    if body.kind is not None:
        e.kind = body.kind
    if body.claimNo is not None:
        e.claim_no = body.claimNo
    if body.project is not None:
        e.project = body.project
    _commit_or_409(db, 'توجد حركة أخرى مطابقة تماماً لهذه القيم — التعديل مرفوض')
    db.refresh(e)
    return CS.entry_json(e)


@router.delete('/{code}/entries/{entry_id}')
def delete_entry(code: str, entry_id: str, db: Session = Depends(get_session)) -> dict:
    row = _get_contractor(db, code)
    e = _get_entry(db, row, entry_id)
    e.deleted_at = dt.datetime.now(dt.timezone.utc)
    db.commit()
    return dict(deleted=True)


# ---------------------------------------------------------------- claims

@router.post('/{code}/claims', status_code=201)
def create_claim(code: str, body: ClaimIn, db: Session = Depends(get_session)) -> dict:
    """مستخلص جديد — the project's guarantee amount is re-derived from claim
    retentions on every claim change (see sync_guarantee_from_claims)."""
    row = _get_contractor(db, code)
    claim = models.ContractorClaim(
        contractor_id=row.id, project=body.project, number=body.number,
        date=_date(body.date), gross_cumulative=body.grossCumulative,
        previous_cumulative=body.previousCumulative,
        retention_rate=body.retentionRate if body.retentionRate is not None
                       else row.default_retention_rate,
        retention_amount=body.retentionAmount, other_deductions=body.otherDeductions,
        net_due=body.netDue, description=body.description, source='manual')
    db.add(claim)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, detail='يوجد مستخلص بنفس المشروع والرقم والتاريخ')
    CS.sync_guarantee_from_claims(db, row, claim.project)
    _commit_or_409(db, 'يوجد مستخلص بنفس المشروع والرقم والتاريخ')
    db.refresh(claim)
    return CS.claim_json(claim)


def _get_claim(db: Session, row: models.Contractor, claim_id: str) -> models.ContractorClaim:
    c = db.query(models.ContractorClaim).filter_by(
        id=claim_id, contractor_id=row.id).filter(
        models.ContractorClaim.deleted_at.is_(None)).one_or_none()
    if c is None:
        raise HTTPException(404, detail='لا يوجد مستخلص بهذا المعرف')
    return c


@router.put('/{code}/claims/{claim_id}')
def update_claim(code: str, claim_id: str, body: ClaimUpdate,
                 db: Session = Depends(get_session)) -> dict:
    row = _get_contractor(db, code)
    c = _get_claim(db, row, claim_id)
    old_project = c.project
    for src, attr in [('project', 'project'), ('number', 'number'),
                      ('grossCumulative', 'gross_cumulative'),
                      ('previousCumulative', 'previous_cumulative'),
                      ('retentionRate', 'retention_rate'),
                      ('retentionAmount', 'retention_amount'),
                      ('otherDeductions', 'other_deductions'),
                      ('netDue', 'net_due'), ('description', 'description')]:
        v = getattr(body, src)
        if v is not None:
            setattr(c, attr, v)
    if body.date is not None:
        c.date = _date(body.date)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, detail='يوجد مستخلص آخر بنفس المشروع والرقم والتاريخ')
    CS.sync_guarantee_from_claims(db, row, c.project)
    if old_project and old_project != c.project:
        CS.sync_guarantee_from_claims(db, row, old_project)
    _commit_or_409(db, 'يوجد مستخلص آخر بنفس المشروع والرقم والتاريخ')
    db.refresh(c)
    return CS.claim_json(c)


@router.delete('/{code}/claims/{claim_id}')
def delete_claim(code: str, claim_id: str, db: Session = Depends(get_session)) -> dict:
    row = _get_contractor(db, code)
    c = _get_claim(db, row, claim_id)
    c.deleted_at = dt.datetime.now(dt.timezone.utc)
    db.flush()
    CS.sync_guarantee_from_claims(db, row, c.project)
    db.commit()
    return dict(deleted=True)


# ---------------------------------------------------------------- guarantees

@router.post('/{code}/guarantees', status_code=201)
def create_guarantee(code: str, body: GuaranteeIn,
                     db: Session = Depends(get_session)) -> dict:
    row = _get_contractor(db, code)
    # unique identity is (contractor, project) INCLUDING soft-deleted rows — a live
    # duplicate is a 409, a soft-deleted one is resurrected in place (an insert would
    # hit the unique constraint and 500).
    exists = db.query(models.ContractorGuarantee).filter_by(
        contractor_id=row.id, project=body.project).one_or_none()
    if exists is not None and exists.deleted_at is None:
        raise HTTPException(409, detail=f'يوجد ضمان بالفعل لمشروع {body.project}')
    g = exists if exists is not None else models.ContractorGuarantee(
        contractor_id=row.id, project=body.project)
    g.deleted_at = None
    g.amount = body.amount or 0.0
    g.retention_rate = (body.retentionRate if body.retentionRate is not None
                        else row.default_retention_rate)
    g.finished_on = _date(body.finishedOn) if body.finishedOn else None
    g.guarantee_days = (body.guaranteeDays if body.guaranteeDays is not None
                        else row.default_guarantee_days)
    g.release_due = _date(body.releaseDue) if body.releaseDue else None
    g.released_on = _date(body.releasedOn) if body.releasedOn else None
    g.notes = body.notes or ''
    if exists is None:
        db.add(g)
    db.commit()
    db.refresh(g)
    return CS.guarantee_json(g)


def _get_guarantee(db: Session, row: models.Contractor,
                   guarantee_id: str) -> models.ContractorGuarantee:
    g = db.query(models.ContractorGuarantee).filter_by(
        id=guarantee_id, contractor_id=row.id).filter(
        models.ContractorGuarantee.deleted_at.is_(None)).one_or_none()
    if g is None:
        raise HTTPException(404, detail='لا يوجد ضمان بهذا المعرف')
    return g


@router.put('/{code}/guarantees/{guarantee_id}')
def update_guarantee(code: str, guarantee_id: str, body: GuaranteeUpdate,
                     db: Session = Depends(get_session)) -> dict:
    """تعديل ضمان — an explicit amount set here wins until the next claim change
    re-derives it (see sync_guarantee_from_claims)."""
    row = _get_contractor(db, code)
    g = _get_guarantee(db, row, guarantee_id)
    if body.amount is not None:
        g.amount = body.amount
    if body.retentionRate is not None:
        g.retention_rate = body.retentionRate
    if body.finishedOn is not None:
        g.finished_on = _date(body.finishedOn) if body.finishedOn else None
    if body.guaranteeDays is not None:
        g.guarantee_days = body.guaranteeDays
    if body.releaseDue is not None:
        g.release_due = _date(body.releaseDue) if body.releaseDue else None
    if body.releasedOn is not None:
        g.released_on = _date(body.releasedOn) if body.releasedOn else None
    if body.notes is not None:
        g.notes = body.notes
    db.commit()
    db.refresh(g)
    return CS.guarantee_json(g)


@router.delete('/{code}/guarantees/{guarantee_id}')
def delete_guarantee(code: str, guarantee_id: str,
                     db: Session = Depends(get_session)) -> dict:
    row = _get_contractor(db, code)
    g = _get_guarantee(db, row, guarantee_id)
    g.deleted_at = dt.datetime.now(dt.timezone.utc)
    db.commit()
    return dict(deleted=True)
