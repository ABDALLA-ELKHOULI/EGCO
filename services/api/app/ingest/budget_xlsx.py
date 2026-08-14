# -*- coding: utf-8 -*-
"""قارئ تقرير انحراف الموازنة التقديرية.

Reads workbooks like `budget-deviation-2026-07.xlsx` — one sheet per month
(e.g. 'تقرير انحراف شهر يوليو'). Layout (verified on the real file):
  row 1  title: 'موازنة اعمال شهر يوليو لعام 2026 م مشروع سدايم' + serial in col H
  row 2  issued datetime in col H
  row 4  deviation table header
  row 5  month row: actual | planned | deviation | delay fraction
  rows 6–7 cumulative to end of previous / current month
  claims section under 'بيان الايرادات الفعليه', total row 'اجمالى المستخلصات'
  financial notes under 'الملاحظات المالية'

Everything is located by label substrings, never fixed coordinates — the sheets
have merged cells and thousands of phantom empty columns, and row positions can
shift when a note wraps to an extra line.
"""
from __future__ import annotations

import datetime as dt
import re
from typing import Any, List, Optional

MAX_COL = 12           # real content lives in columns A..H; phantom cols reach 16k+

#: Arabic month names → month number (keys stored in normalised form)
_MONTHS = {
    'يناير': 1, 'فبراير': 2, 'مارس': 3, 'ابريل': 4, 'مايو': 5, 'يونيو': 6,
    'يوليو': 7, 'اغسطس': 8, 'سبتمبر': 9, 'اكتوبر': 10, 'نوفمبر': 11, 'ديسمبر': 12,
    # common spelling variants
    'يونيه': 6, 'يوليه': 7,
}


class BudgetParseError(Exception):
    pass


def _norm(s: str) -> str:
    """Normalise Arabic text: unify hamza/yeh forms, drop tatweel, collapse spaces."""
    s = s.replace('ـ', '')                      # tatweel
    s = re.sub('[أإآا]', 'ا', s)
    s = re.sub('[ىئي]', 'ي', s)
    s = s.replace('ؤ', 'و').replace('ة', 'ه')
    return re.sub(r'\s+', ' ', s).strip()


def _num(v: Any) -> Optional[float]:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _row_numbers(row) -> List[float]:
    """Numeric cells in a row, left to right, skipping column A (the serial م column)."""
    return [n for c in row if c.column > 1 for n in [_num(c.value)] if n is not None]


def _row_text(row) -> str:
    return ' '.join(str(c.value).strip() for c in row
                    if isinstance(c.value, str) and c.value.strip())


def _to_date(v: Any) -> Optional[dt.date]:
    """datetime → date; strings like '00/07/2026' (day zero) are unparseable → None."""
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        m = re.match(r'\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$', v)
        if m:
            try:
                return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                return None
    return None


def _title_month_project(title: str):
    """Extract (month_date, project) from the sheet title line."""
    norm = _norm(title)
    month_no = None
    for name, no in _MONTHS.items():
        if name in norm:
            month_no = no
            break
    ym = re.search(r'(\d{4})', title)
    month = dt.date(int(ym.group(1)), month_no, 1) if (month_no and ym) else None
    project = ''
    pm = re.search(r'مشروع\s+(.+)', title)
    if pm:
        project = re.sub(r'[\sــ]+$', '', pm.group(1)).strip()
    return month, project


def _parse_sheet(ws) -> Optional[dict]:
    """Parse one sheet; return the sheet dict, or None if it is not a budget sheet."""
    issues: List[dict] = []

    # -- title: first non-empty string cell in the first rows
    title = ''
    serial = ''
    issued_on: Optional[dt.date] = None
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=MAX_COL))
    for row in rows[:3]:
        for c in row:
            if isinstance(c.value, str) and 'موازنه' in _norm(c.value) and not title:
                title = c.value.strip()
    if not title and 'انحراف' not in _norm(ws.title):
        return None

    for row in rows[:3]:
        for c in row:
            if isinstance(c.value, str) and c.value.strip().upper().startswith('EGCO'):
                serial = c.value.strip()
            if issued_on is None and isinstance(c.value, (dt.date, dt.datetime)):
                issued_on = _to_date(c.value)

    month, project = _title_month_project(title) if title else (None, '')
    if month is None or not project:
        issues.append(dict(severity='warning',
                           message=f'تعذر استخراج الشهر أو المشروع من العنوان: {title!r}'))
        if month is None:
            return None

    out = dict(project=project, month=month, serial=serial, issued_on=issued_on,
               actual_month=0.0, planned_month=0.0, deviation_month=0.0,
               cum_actual=0.0, cum_planned=0.0,
               cum_prev_actual=0.0, cum_prev_planned=0.0,
               delay_pct=None, completion_pct=None,
               claims=[], notes='', issues=issues)

    # -- deviation table: locate the three data rows by their labels
    def fill(nums: List[float], actual_key: str, planned_key: str,
             dev_key: Optional[str] = None) -> None:
        if len(nums) >= 1:
            out[actual_key] = nums[0]
        if len(nums) >= 2:
            out[planned_key] = nums[1]
        if dev_key and len(nums) >= 3:
            out[dev_key] = nums[2]

    claims_start = None
    notes_start = None
    for i, row in enumerate(rows):
        text = _norm(_row_text(row))
        if not text:
            continue
        nums = _row_numbers(row)
        if 'حجم العمل الفعلي شهر' in text:
            fill(nums, 'actual_month', 'planned_month', 'deviation_month')
            if len(nums) >= 4:
                # the delay column carries a negative fraction; store magnitude
                out['delay_pct'] = abs(nums[3])
        elif 'التراكمي حتي نهايه' in text:
            fill(nums, 'cum_prev_actual', 'cum_prev_planned')
        elif 'التراكمي بنهايه شهر' in text:
            fill(nums, 'cum_actual', 'cum_planned')
        elif 'الايرادات الفعليه' in text and claims_start is None:
            claims_start = i
        elif 'الملاحظات الماليه' in text and notes_start is None:
            notes_start = i

    # -- claims: rows between the revenue header and the 'اجمالى' total row
    if claims_start is not None:
        for row in rows[claims_start + 1:]:
            text = _row_text(row)
            ntext = _norm(text)
            if 'اجمال' in ntext:
                break
            m = re.search(r'المستخلص رقم\s*\(?\s*(\d+)\s*\)?', text)
            if not m:
                continue
            nums = _row_numbers(row)
            date_cell = next((c.value for c in row
                              if isinstance(c.value, (str, dt.date, dt.datetime))
                              and c.column > 2), None)
            claim_date = _to_date(date_cell)
            if claim_date is None and isinstance(date_cell, str):
                issues.append(dict(severity='info',
                                   message=f'تاريخ مستخلص غير مقروء: {date_cell!r}'))
            # zero-amount claims are kept — a claim planned but not yet issued
            out['claims'].append(dict(no=m.group(1),
                                      amount=nums[0] if nums else 0.0,
                                      date=claim_date))
    else:
        issues.append(dict(severity='warning', message='لم يُعثر على قسم الايرادات الفعليه'))

    # -- financial notes: free-text lines until the signature block
    if notes_start is not None:
        lines: List[str] = []
        for row in rows[notes_start + 1:]:
            text = _row_text(row)
            if not text:
                continue
            if 'المدير المالي' in _norm(text):
                break
            lines.append(text)
        out['notes'] = '\n'.join(lines)
        cm = re.search(r'نسبة انجاز[^%]*?([\d.]+)\s*%', out['notes'])
        if cm:
            out['completion_pct'] = float(cm.group(1)) / 100.0

    return out


def parse(path: str) -> List[dict]:
    """Return one sheet dict per budget sheet in the workbook.

    Non-matching sheets are skipped (noted as issues on the result), never raised;
    BudgetParseError is raised only when no sheet matches at all.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise BudgetParseError(f'تعذر فتح الملف: {e}') from e

    sheets: List[dict] = []
    skipped: List[str] = []
    for ws in wb.worksheets:
        try:
            parsed = _parse_sheet(ws)
        except Exception as e:                     # a broken sheet must not kill the rest
            skipped.append(ws.title)
            continue
        if parsed is None:
            skipped.append(ws.title)
            continue
        sheets.append(parsed)

    if not sheets:
        raise BudgetParseError('لم يُعثر على أي ورقة تقرير انحراف موازنة في الملف')

    # note skipped sheets once, on the first parsed sheet
    for name in skipped:
        sheets[0]['issues'].append(dict(severity='info',
                                        message=f'تم تجاهل ورقة لا تطابق الصيغة: {name}'))
    return sheets
