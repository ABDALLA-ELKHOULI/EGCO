# -*- coding: utf-8 -*-
"""تصدير Excel — GET /api/v1/reports/export.xlsx.

Three sheets, right-to-left, Arabic titles: الملخص (summary), الفترات (periods),
الموردون (suppliers).
"""
from __future__ import annotations

import base64
import io
import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

NUM_FMT = '#,##0.00'

# شعار الشركة — نسخة مصغّرة من build/icon.png (بلا الشريط السفلي «Emaar Gulf
# for Construction» غير المقروء بهذا الحجم)، مُضمَّنة كـ base64 داخل الكود
# بدل مسار ملف: الخادم يُشحن كحزمة PyInstaller (onedir) لا تحمل أصولاً غير
# البرمجيات المُدرجة صراحةً في egco-api.spec، وأي مسار نسبي للملف الأصلي
# لن يوجد على جهاز المستخدم. التضمين يضمن ظهور الشعار في كل بيئة تشغيل.
_LOGO_PNG_B64 = (
    'iVBORw0KGgoAAAANSUhEUgAAAKAAAABbCAYAAAD0mo73AAAcpklEQVR42u2deXRcV53nP/fd92ovlfbVtuTdsh3Li2wnJmsTSEgmCSFMwnaAADPTEHqgSRg43dCTTBM40NPdZIbDNoGGgZBMSAgN3SEsCc5mJ5b32I5tyYtsydpLW6m2996980eVtUR2YsuSHSfvd04dLVWqurrv+37793eFnU1pPPHkAonhbYEnF1LMk98IYSCEh0dPZl60VmitJgPQkKa3O57MuCjXGQWgp/I88XxATzwAeuKJB0BPPAB64okHQE88AHriiQdAT3j7VEI8mboorXGS/dhDnSQGjxGbvZZAuMTbGA+AzEAZCZxUP06yn3RvC8n4AdLdh8n0t0K6Dy3DRG7/P95GeQA8U0DpfCkSQEx8DlDZFNlknEz8MNm+Y6T6DpLs2INykmhpEQhXEq6op3DRDWSGDpFo3YIv5Gk/D4BnKEKMgc5x0rgjvaTjx8nEj5Ds3k+27zB2ug+BhQgXEyyZTWHDfyRYvpRQcS1mMIYhwM0m6Xi+CSNQiOkLeMh6uwNQaw0CxGu02qjfplxUqo/kYBeZ/iNkT+wj3bcfe7gXpTVmKIZCE561nsq6ywkU1yHDMQxpjb5jdqSPoZanSTQ/x0jXK9jZBMVr/9NpPtGTt5cGFGPQU8rGzSTJDvWS7msh1bWDTPdh7MQJXHsIOZJAli4iOv8aAlXL8RfNwYqUED/wDMGSOqLVy0e7ONK9h0ic2MnwkWfJdB/A0GCU1hNbfivRuksJFs/1UPV2B6BrZ8gOdZLuP8ZI5x4y8QM4Q51kM4MoM0gkWoOvYhGFK+8g1X0AJzOA6Y+SkRIjm8JnGJi+ENIwUOkEyc79DB5/mZHDm3Hih8AfxFezltiyRRhF8yhbcj3S66W8wADUObMnjOk1QHlLmtdqpwabnYqT7T1Mqq+ZdE8L2b7j2KkutGEgg2XIeAuibBnVaz9FpGYVvkjJ6Ft1ZQcIFa6jePZq0gPtpDoP0LXpQZRWyHQ3qf7jaNvBFwoTqFlLZN1HCZYtxh8pIdl/lNRAxzjw6VMv0pOZBWAmMYBhWliB8ExY0jEwKpfsSC/ZvsOkeg+R7D6KHmghk2hFqCAiVomvaDaR+qsJVCwlGKsDZ4Su3Y/jL57HSOd2Ej0HiVbWE6ppIBAsxhQhpK3IDvWR6jpAqm07qZ7t6EQv/rJVFC29ifCcSwmUzscKRCauTbkYuON/4yHqfANwsPlPdLzwz7hWAbGF11K84F0Ei+dMAM7rtWYLRE5vvOYPlFY4iT7ckW6S8eOkT7xCJn4Qe6gT5TgY4Ri+wjocvw9/aBXaDFNYfwOx2jVI0zf6PtmB4xAop3LFbSg3zUhPC4kTe+k79B3CxTW4iQ5S7XtRqREcoQiXLqa44SNEZq0kWDQX+Xpd4lrnlJ4n5xeAWgPKpnf7z+jY9hAF9Tfi9xcxtO/fie98jLKVH6Fy3YffEIQnOShaOzipYZyRfpK9B8l0HyTbc5B0fztKZTCcQTLKIbbkZirX3IkVq8SKlGEFopzY8TCRmkYsoPvQ0/Qf/jNFlQ2EalfjL6jCMCTCzaCVzgEXA9NViM49xA/+BlE4h3DpKgrq1hOsXkEgVnnmekxotFAeis43AAWazpcepG/3E1Rd9UVKl16PAMoabqN36w9Jtm5ENX4IKcVpgoQkdiKO3X+Mke79pDp3kR48js7YGAisWBW+0sWULLqOQNk8+lv+QDRcRXawExEqIlQ6b/S9LC0xXJdg1RJqyxeSGe5i6OhmurY+jBWbTbR8LnqgnePP/iPO8SZS6R5kuIxI3aWUz/1bAqV1WKGYZzwvFgBqDWhFpr2J0saPUbb0+tG8mxksIFy5kqGuI5AHn+tkcRN9pAeOM9Kzn1TPAVR/K26iBzfVhwzXEKldS2jOpQTL6vFHK7GipRMIUsnjLxGtewcC6Nr1OENHt1Gx8hasYBRMA6Xs0df6IuXEai9FKIfeLQ8St1OYZhCK6ogtfg+lc9YSLJ2HZQW9q3/RakABSIkVKppQUVCug20n0Olu+nb+lnTfq2T69qMGO1HKQISi+AvriC66gUz8ANosIJtNI0uWEZmznHDpglOH1yi0sgmVzGPOFZ8h/upTHHv+f1HRcBvC8mE7CZLxI4y0vMhg20s4fYfAkASr1hCqvZRg9QqChTUY0vQ03cUAwFx5VCOEyKVWxjlzuW8lQmvsdIJU72FGel4l23uUVEcT9kA3RqaPnm3fJVi+nGjNpfhXLcZfXIcvWo70BRFAx46Hic5aSzBayeCRl+jc/QSGYRGrWUNk1kp8wego/FBjvr4hfZQuvxl/8Ww6XvwOMpskkxlCpYcwrQL8sxsoWvY+ItUr8EVKJ0bRWnNG0ZEnFxaAYlz+Q4hclOpmhnGGe0j1HiLd20xmsIPU9p8T3/ZjMCUyWkugaj3+kn6Gjj6LVb2a4qXvpbB27am1qGujlY0MRCiuv5aiRVeT7GlmsPUlBo+9jL9oNrE5awiVLkBLicLETnST7HiFobYmkse3oJNJKKgiWncFBXVXECpfgAwWnFbLCQ98F4kJFmCn+kme2Eu6r5mR7oNk40dwsoMY+DHCMSw3ibZi1Fz7LYLFc5CBAgxpkGj+HamBDqrXf5L4K08Sb91MxfJbCBfXngLm4zSrNAlX1hOurMdN9TPQvpveXb9CmiYMtNPduh23vx3DSSBLFlCw9DYiNWsIls7FtPzelXyrmeCOzd9n4OBv8UcW4yucRaz+JkJVywgUz8GJH6H9ya9glS1jpHMX0VkrTtptcDWW6xIsmsesKz/DYNs2OnY/TCS6gOLFf4EvUpwzpYb/lI0Crp0mmxiAoTi67ygDnS9ihWrxl9cTXfcxorWX4o+Uvn5+zpOL3QRr9MggRZd8lJoNf8n46pqbGaHzwDMgfRTNv4qRVDddO56gcvWtuQQyOpffUy5Ig8LZaymoWkr80HO0b/4xkaoVlCy5GiUF2sjlAe1UnFRfK4kjz5Noa8Ie6cbwFRGuaaS68YOEypfgj5R5AcTbRgMKgQwEGN7/JD2FdZQsvHq0x61z9y/xVzcgksdxlENVw+20b3yAeMtmihdcBoZAi5NAzn1jmGFKF7+HwtmXMbD/aY5tfhA3eQKdTTK463ESxzahs0lkdBYF867BP2cVoZJ5+AMx7wq9XYOQsks/g/D/nL5N/5vBXY9QsuYjWIEYmUyG2tVX07b/CUQ+Kq287E5aX/gOVrgYwwwgtD51K7udxghGcPc9Rrp3HyMF1QQLFxJbcTvR2Y0Ei+qQXjOnB0CtNf5oOTVXfoHiFbcT3/Eo3X9+AJdhqi77awQmyjBz+RHACpVS0/hROrb/goLCarTpAylx7BR23xGGj75Mqm0bmb7DONIiWHMJZWs+QbRqGf5o+WjS2Survv1EnJyQahjm645nS/Y007fjpwwfacJXvgSjv5PoOz5B2eLrRl+T6N5H29P3I5QiOucyhlubcIY6keESQrPWEJ3TSKh6Bb5wyVvCn0v1NZMa6qB47pUekji78WxKOWdXCQmVLST07q+R6NxHfNuDJEeOMvjK7/CFKwhEKkh172GgfRtGohtlZ0i27yQ690qCtWuJlM7DChR4O+/J1DXgRD6Fw/F//Swj/e3IrAZT4ooM/uKl+Px+7JE4s295AOst7s95GvA8asDxlVkhDIQWlDf+JVa0kJ5NP6Dimi8Rq15GsuVpenc+gTmuJ88TT6a9HQty7fexue/ATQ8w3Pw0hdXLUG4GrfXEVnpPPGG6ZsOMhqpqtK5ftPg6cGziB54F00QLL571ZKaHE+lxdVzDpGrtB+lvfprMQDdCSC+p4skMAVDksCUQuOMMuFlQTdny6+jb+yu0Ib2d9WRmSUkCGD6+g6yjMBwnZ3YNCxMDMa5D2RNPpt8EC41Akjz6IlJJfOFSfKEy/KEKwvPfifbCD09mTAPqnP5TBhQuvJFMsoOypdcipQVAQsVJd+zxOLKezGwQonSacPlcQtEyTux4fAJ53MOeJzMOQI1AKU1Z/Y0YI50MHHpxzDv0AmBPZhqAAonAAGFQvuZD9Dc/QyreijDw8oCenIfRHOMIPlaknPIVt9C97WfEqurRQoJhnMuco7O24ul0hhc2beblLU0MDg1NGwHJH/CzfOlSrrrycirKys5wLWle3Pxybi2DgzPGxBMKZs2q4lOfvJNAYHLdvb+/n43PvcCOXbtJZ9KnnZV4tlcnGo2yetVKrthwGdFo9MIAUKMmgCxcvYJUfyvd236BDBWTGe7KPz1uforIJ7CFHkOZfk1yx7CwgjHEWeQSW1qOcN/ff4NNL2/BdpxpA19uebmFzq2dzRe/8F+58YbrX/dvDhxs4X/c/01e3rIV27ZnlAbqZm0aGpby0Y9+ZBIAt2zdyr333c++/S1o9OgMHqZpFJqUv2DFsnr+5st3s25t4/kHoNQOqf5DDHU2o1QGkFixeZiWj3TvVlqf+E8I286zRPJXUwmUYWAKUFjg2gjhoDAQaAwNTqiUuhu/Saiw5ozW0d7ewWc//2Ve3b+fYDCEzzfGj5rOCV3H27q5+0tfRaP5Dze925SvO9p6jLs+dw8tza0Egn4sMzCj7rArMwT9wUk33Padu/js5++hp6ePYDAyjus9nZsCu17Zz12fv4cffPcBVjc0nF8AGkaIgf1PomQhkaolaK0xhEV4/jVktryKDJdTsuhmUC4KlVd1BtlMnPThTWgnS2TlLZjSRJ2ckpUZYGDPLzGczBmtwXEcvvUP32bfqwcIh8OjTRDMwGR8yzJx3Cz3f+N/csmyZdTWzpnwmkwmy9e/+U80txwiFIqcl8FZCgFaMt7lHh5O8PX7/5Hunl6C/uCM7MdJPykQ8NPb28t9997PT3/8QwqLCs8XADWuVpQ1fJhMsp9ApJxg8SwAkulOBioaCM5/N2k7SeXq2zHG3aGZweMMBIqwhCQx1EvZhjsx8qbcTfUx1PzkGa9i167d/PnpjQSDgdFp9zMnGtM06TjRzS8ffYJ7vvi5Cc9u3bqNZ597gUAwmF/LhclFvfD8i+zatYdAIDLje6K1xu/388or+/ndk3/ggx++/XxFwQKhbHyhYkrr30XX1oewk/15WrCLhY+ahlvx+Xwce/572OnhcXnCLFrbFK+4CX8kRuuL30M52ZNODUK7Z9zIsPnlJoZSw9M+lfW0W6UNpOlj4/ObSKYyCMMYdeyfff4F0pksAnnewGfokyPixgFw00vYSuWbRWZ6bLAAbaCFyZ82PofjOOcxCBEGSimCpQspnLeOrqZfUn35J/OT6TUag9JLbkUe3Ejrc99h9vpP4I9V5Ct5uecrVt5GV9PPOLr5u9RuuCvf6Hrmd23b8Q4QchJeFdlcAKDODQxSmpiWOc6BEkjTpLu3l3h/nLKQGE17Hjl6LNcFpCdriaydnRF77Nppsk5y7GfXpa29HSGNSaVTpezcnuhzI/RbPt84i5b7Kk1J6/Fj9Pf3U3aGmYJpSMPoUU1VuOAaMoleul75LZFIJA+KXPxYtOhqZCDKsc0PUtP4IUzLAq1Gg+OSNR9Ebf05bZv+heqGG0EYZ1xLdhwF2ngNXUARi4ZZtHgBhmHmNdLZmxbDMDjR3sGxthMIOS65LjRKKZTrjv4qRzlVk9IcWmnC4QCN9Q1IY/q1kZ1JsmBhLaaUo+tWLgj9mmmzrqK8tJi582pBm1POEijlcuDAIQaHRzDGWx0B2tUopc+nBhxrCdRA+Yr30rblh/ScOA5SgpSjl6NgzhpkKEbfjscIzl6J9kVGTyYyhUXVujvpbvoZ7U3/F84CMKfaRsd2uGTpar7//QcwDCOfRplCdGZIHv1/T/CVe+/HNM3RDxOuPnVmRYhJK3IdhwVz5/CjHzyAaZpTXsvrBQJCiNz68oMADGGdAqg2G9at5+vfuHfKaxAIXNfl05/+AhtfeAmf33eOmdtz1YA614g6CgTDorrxExz//dew+w7RvvE7iNEB3i6uYaCzPfQ+cz9G6ULsviMI7earJhJX2CQP/xEjmz2n3JkGpDQwTXM0uJmyubFMUqkUUpnjtIDGMiVKnZmellJimuYoSGZe9KnvVME5r8EwDKQp3wyVkNzQyJHWHWgng4tEaAMtNTJcTJIsYqQHA3ts1gwCIXz4518DysFOdI17LwOESXDOOxHSQvii53QBlJ6eIKB+2SK+ePddo50+OTPnYlqSwsIYWvW8oZV4M0yQ1gjUNCXE9TSX+s2pdkRL4WP42LM4UhGKVeMqgStd3FQ/vqJF1L3nv0/J7xFvosHz9fX11NfXc3paZhcXi2jeav2AGopW3oEjfRQtfg++cC4JOdjyFD27/g1jqqw44XVz4XXDnImmcvBFKimetZbOHQ/jZFN5J93F1F43jCczflBNjvsRKl9AbOAI3bt/Q9WaOxBCXlTtgEeOtvLS5pcx5JnlDLVWWD6T6971LqQ38vfCAVDnk80AsUXXktr+CD17f084cHFNot+xfSf3fPmr+PxB9BkEL0opYgUh1qxeTVXMO6DwwmlArSZUCCpW3Ern5h/RmzyBMM2LhhcspUkwGMDynTkA/f5APsfoyQUlprvjUkLCDFC+/mPo1BB6CjXBCxkeanWyhKbP8OHJBU/DIF3SbXvpM/xoN5ur75oBzFgZdrz7ItqC81G092RaZ8NoNMIIkWj5EyMdezCkHyFMhBYYMpTjilxUGTJPq11k07EEWmUoWnMHmZEBQmWLCeT7AYcsh77uwxcNL1gLhdYKpZlYJ80fquQdbvOm5QVrZEEl5YvfRWfTz3AyubYg7V5cYzkEGstn4PP78PvGHgG//zzWbz0NOIULp5AOBKtXUDjURseWf2HWO/5LrhVLGBeNWbv8ist55KEfY5q+CT0QUko2bnyBf/r295CWB8Q3IQDlaNdKbPH1JBM/pWvP7wmHffmm0ovDdJUUF1NSXHzK55oPtkx/C5Un526Cc53gY/1vGoPKVR8mM3iY/sOb0dJ6S1w2V7lvEfjpaSVlXlAA5hxziZIGKt88aojcgTVVaz9EdrgTrVyEYVzYzMpbvo/krEJGxJuiMewsTLCd7OfElp8g3BQoJ99mL9CGBvzY8RMM7f8dya6dCNdFC4Fr+DBwyQ620fbnf0C4mXy5brxJ1mjDAmFhqBRoF52b55F7nWEAEq2zGHrifau1IjxnAyVL3jmO0y5O2zw5HawvrdU5BzkXtBn1ZMflNAwNNQwjX/3SMw9ANz1Iev8TuK6Nv3od2h3OARCFFn60m8Du2YM9fBQrUolSWYQIoDID6GQvIy1/JFC2ENvNdRSfnDCgDAeR6CU7MoC/chlSGyjXQRtOrpEh2Y891IWvYgWIDErnAh6Bhdu9D2GYlCx552lzeKZp0tLayk9//hChQCjPL9FnDxut+OMf/4R5SiCfin+sT9HWLzjR3sGPf/IQ0Wh42mmSjm1TWlLMNddchWXl3B6ls0wgCqORpsX+5sM88uhjUwaiEIJUMsPhQ0cwX9sVrY38qVnTyAkRgAyVYZWvIli1nIpLbso1IOTvgPbffI5A5XqSQlG+5N2EYtVoBEMHn6Jz12NE6q7CilRQuvgvEMIY1SRCaAYOPkN2qBstQ5SvfB9C5DqNhZD0v/oUqZEeHBQ1DbchrXDurhOCE0//PVqNAUIaGoQzqQW+/UQP9933zRyQxFRncGpMaSAt/6RtNYSBlHLiXsnJnA/DMOjp6eNrX/vW6Mmg0wrAbIqVDZewYcOlWJaFEAIhxaSbwzIt9r7azN/87X1oYZ4DDVThsyxMab5moorGkMaUwH361Wg312TQ+AHi+/9A1+5fU7nytnwS0MFwHaxYJZVldXTvfJzqy+7E8udOK7eMANUr30/3rl/Rvu1Rala/H2n68xuj0K5DySW3EH/1KXq2PU7l2tvAyLe9qwRFdRvI9h+hc9vDVK39ONLy52kNYgIHtrq6EqFF7lT3cbsuhYEVCJ9TBCtGAywmMczKikspLCqCdBJ0jqQ0p2YWvIYVpoVASBN/0JwRT1IaAr8vOOHmq66oQk9ip2lMw0QEwue0DjEupJmwJ45LVWUlsVjB9AUhSrkIpZFmkFnr70TFD9G167G8FjNyy3FdgmVLKKpZRsfWh1DKzQUkWoMZpLLxQwRCEY4++22yIwO5rI0GgYEQBhWr3o9Bhq4tj6CVGkvvaEXR4ncTLK7lRNNP0I497qKObeHaxkYCvvAkOqAYhZ6Y8uPUno7AsW3WrllBJJSbxnDyNe/YsBbfpHyhyMef4pzWcvo1ikleyPp1jafhFepzXsfY/zLGwMZwUTrDlZdfhs/nm94oOBfJaoQVpPryu7B799O998lcolmYnDzJumDhtQSilbRv+0Werpl3VoVB+fJbKKpdz7Hn/pmRnubce+Z5wwhJ+doPYpOhbetD49tTcjm6xdcTilVy7OUf4To2wjAnENdXr1nF+vWNpDPp85J2VMqluDTKHR943yS/77IN62hsbCCbzZy38t2pPuXKq69gwaL5ORL6jC9AY9sZ5s2r4+abb5zeNIzQOheZ5k2j8IWpuepusr3NdL76FBg+lDHGFitf8V58WhPf+wxCBib4PMULrqZq9QfoanqEoba9KMsPJ5lmQjJr/cfxKU33zt+ikWN0TwGly24hWFLFiR2Pgp2Y4GAHAn7uvvsuKstKsdPpGZuEJoRAKYVtJ/mruz5F/dKlk14TCoX4b1/8K0pLC84bCMVJUrgYn1gv4u6/vgufZeI49oytQwiBYzv4TT9fuvsLlJeXT82N+LuvfuXek2Z1fO7OTvQytP93WKULcNNDZIc7cRJ9mNESBnc/Qqp7H1ZBFUKaZIe7yAx2IEMFjBx8EtcdwRebjZ3oITvURWawHaEFpt9Hz5YfkE0PYfqjZIbbsYc7yQ514YsUM7jv3xjs3IMvWo7KpsgMdZIZOIrfV0i6bTNDzU8TKG8gNvey0XVWlJdxybIl7N65j66entzUAqVQ7jQ8lMJ1FdlslkjYxz2f/zSf+PidoxfVScVxMgmCRbUAVFZUsnzpErbv2EVXVy9Ka7Q7TWs5xcOxbcrLSnnfrTfhH0cUnz9/LrWzqtm2dScDw0Mopad5T1wymSzVFSXc93df5oYbrjvL9JYaC0pPd1rmSG8LR//9SwgnPTpTMpeXM5BCo8mikShUbiJq7hRDJAKhHBSgjLz/oHVOcRkmkEK64BghIJcy0FpiCDMXZLiZXDpG5DpVUAJTg2sa6Eya2LL3MeuKT0/6pzo6unj8V/9KU1MTicR0TQMFy5Isrl/IzTfdyOpVDWd0WmZbewe//vVvaGrazkgiNXOVGifLwkXzufferxAKBSc9f+BAC4//6gle2b0X23anJUOvtSYY8rFyZQO33noT8+fNPafTMk8LQOXaZEbiCG1PalMaG0WRA5YetQUne5h0Li0kTtXJOpadPxlRiQmfMDFvN+EnDdIfxRcsfJ1/zsVxpmdSv9ZjUxaYwnGtSqkpTYw6W1NoWdYbzNCxUe40nR6pc8OIxqehZuS4ViEtggUVF13hyZASnzw/R4VpPUbMOl3lYCqR4bSXu0zrnPiP570Up0/a6XH5MC4Knr84f58pBBoH96T2PmXJTrwJZh6cxz1BnHXQM8kEO6kh2jf9EJ3pyaVLPDntdXUzCVwnjT9cfs41Y94S3UM2seV3UFrbOHUTrOwkqeYnIdGVy/V58jomNhdkpZXyugYRuHaaQPlaqG2cugk2rCD+2sshO5SveHhyWmq+m4sscxNJhWcRbJdAyfxzM8EnR70Kb0rQG7MDtZt3B6W3V+Mn7QnjdbfjDaNgKT3fb6Z5/Z7k5P8D6Sne0j60DKAAAAAASUVORK5CYII='
)


def _logo_image() -> Optional[XLImage]:
    """كل ورقة تحتاج نسخة Image مستقلة — openpyxl لا يسمح بإضافة نفس الكائن
    لأكثر من ورقة.

    تُعيد None إن تعذّر إنشاء الصورة (Pillow غائبة عن الحزمة مثلاً). الشعار
    زينة، والتصدير وثيقة يحتاجها المستخدم: انهيار التصدير كله لأن صورة لم
    تُحمّل مقايضة خاسرة. الفشل يُسجَّل ولا يُبتلع.
    """
    try:
        img = XLImage(io.BytesIO(base64.b64decode(_LOGO_PNG_B64)))
    except Exception as e:                      # pragma: no cover - يعتمد على البيئة
        logger.warning('تعذّر إدراج الشعار في ملف Excel: %s', e)
        return None
    img.width = 88
    img.height = 50
    return img


def _add_logo(ws, anchor: str) -> None:
    """يضيف الشعار إن أمكن — وإلا يمضي التصدير بلا شعار."""
    img = _logo_image()
    if img is not None:
        ws.add_image(img, anchor)


def _style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = Font(bold=True)


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 40)


def _contractors_sheet(wb: Workbook, contractors: dict, first: bool = False):
    """ورقة المقاولون — no ageing columns: their ledger has no due dates."""
    ws = wb.active if first else wb.create_sheet('المقاولون')
    if first:
        ws.title = 'المقاولون'
    ws.sheet_view.rightToLeft = True
    ws.append(['كود المقاول', 'الاسم', 'المشاريع', 'المحمّل عليه', 'المدفوع',
               'خصومات وتحميلات', 'المستحق له', 'الرصيد'])
    _style_header(ws)
    for c in contractors.get('rows', []):
        ws.append([c.get('code', ''), c.get('name', ''),
                   '، '.join(c.get('projects') or []),
                   c.get('invoiced', 0), c.get('paid', 0), c.get('deductions', 0),
                   c.get('outstanding', 0), c.get('balance', 0)])
    t = contractors.get('totals') or {}
    if t:
        ws.append(['الإجمالي', '', '', t.get('invoiced', 0), t.get('paid', 0),
                   t.get('deductions', 0), t.get('outstanding', 0), t.get('balance', 0)])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    for r in range(2, ws.max_row + 1):
        for c in (4, 5, 6, 7, 8):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUM_FMT
    _autosize(ws)
    return ws


def _priorities_sheet(wb: Workbook, priorities: dict):
    """ورقة أولويات السداد — نفس القائمة الحتمية المعروضة في قسم التقرير التحليلي
    (F.build_priorities)، بلا عمود «ضمن الميزانية» — ذاك تقدير محلي في الواجهة
    فقط، فلا يُطبع كحقيقة من الخادم على وثيقة مُصدَّرة.
    """
    ws = wb.create_sheet('أولويات السداد')
    ws.sheet_view.rightToLeft = True
    ws.append(['#', 'الاسم', 'النوع', 'المبلغ (ر.س)', 'السبب'])
    _style_header(ws)
    for i, it in enumerate(priorities.get('items') or [], start=1):
        ws.append([i, it.get('name', ''),
                   'مقاول' if it.get('partyKind') == 'contractor' else 'مورد',
                   it.get('amount', 0), it.get('reason', '')])
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=4)
        if isinstance(cell.value, (int, float)):
            cell.number_format = NUM_FMT
    _autosize(ws)
    return ws


def build_project_summary_workbook(payload: dict) -> bytes:
    """ورقة واحدة — ملخّص المشروع بسطر واحد لكل شركة، بنفس آلية build_workbook.

    عمود «أقصى تأخر» و«شريحته» يُتركان فارغين للمقاولين — لا يُخترع صفر لتأخر
    لا معنى محاسبياً له (انظر تعليق `delay` في report_service.project_summary).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'ملخص المشروع'
    ws.sheet_view.rightToLeft = True
    ws.append(['اسم الشركة', 'رقم الحساب', 'نوع الطرف', 'إجمالي المفوتر', 'المسدد',
               'المتبقي', 'المتأخر', 'أقصى تأخر (يوم)', 'آخر دفعة — التاريخ',
               'آخر دفعة — المبلغ'])
    _style_header(ws)
    for r in payload.get('rows', []):
        delay = r.get('delay') or {}
        lp = r.get('lastPayment') or {}
        ws.append([
            r.get('name', ''), r.get('account', ''),
            'مقاول' if r.get('partyKind') == 'contractor' else 'مورد',
            r.get('totalInvoiced', 0), r.get('totalPaid', 0), r.get('outstanding', 0),
            delay.get('amount', '') if r.get('delay') is not None else '—',
            delay.get('days', '') if r.get('delay') is not None else '—',
            lp.get('date', ''), lp.get('amount', ''),
        ])
    t = payload.get('totals') or {}
    if t:
        ws.append(['الإجمالي', '', '', t.get('totalInvoiced', 0), t.get('totalPaid', 0),
                   t.get('outstanding', 0), t.get('delayedAmount', 0),
                   t.get('maxDelayDays', 0), '', ''])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    for r in range(2, ws.max_row + 1):
        for c in (4, 5, 6, 7, 8, 10):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUM_FMT
    _autosize(ws)
    _add_logo(ws, f'{get_column_letter(ws.max_column + 2)}1')
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_workbook(analysis: dict, periodic: Optional[dict] = None,
                   suppliers_rows: Optional[list] = None,
                   contractors_only: bool = False,
                   priorities: Optional[dict] = None) -> bytes:
    wb = Workbook()

    if contractors_only:
        # تقرير مقاول واحد — his sheet and nothing else; the supplier sheets would be
        # empty and would read as "no debts" rather than "not in scope".
        ws0 = _contractors_sheet(wb, analysis.get('contractors') or {}, first=True)
        _add_logo(ws0, f'{get_column_letter(ws0.max_column + 2)}1')
        if priorities is not None:
            _priorities_sheet(wb, priorities)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ---- الملخص
    ws1 = wb.active
    ws1.title = 'الملخص'
    ws1.sheet_view.rightToLeft = True
    meta = analysis['meta']
    summary = analysis['summary']
    rows = [
        ('الشركة', meta.get('company', '')),
        ('الفترة', meta.get('period', '')),
        ('رصيد أول المدة', meta.get('opening_balance', 0)),
        ('رصيد آخر المدة', meta.get('closing_balance', summary.get('outstanding', 0))),
        ('إجمالي الفواتير', summary.get('total_invoiced', 0)),
        ('إجمالي المسدد', summary.get('total_paid', 0)),
        ('المديونية القائمة', summary.get('outstanding', 0)),
        ('المتأخر', summary.get('overdue', 0)),
        ('مستحق خلال ٧ أيام', summary.get('due_within_7', 0)),
        ('عدد الموردين', summary.get('supplier_count', 0)),
    ]
    if analysis.get('contractors') is not None:
        rows += [('النطاق', meta.get('scope_label', '')),
                 ('عدد المقاولين', summary.get('contractor_count', 0)),
                 ('رصيد المقاولين', summary.get('contractor_balance', 0))]
    ws1.append(['البند', 'القيمة'])
    _style_header(ws1)
    for label, value in rows:
        ws1.append([label, value])
    for r in range(2, ws1.max_row + 1):
        cell = ws1.cell(row=r, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = NUM_FMT
    _autosize(ws1)
    _add_logo(ws1, 'D1')

    # ---- الفترات
    ws2 = wb.create_sheet('الفترات')
    ws2.sheet_view.rightToLeft = True
    headers2 = ['الفترة', 'من', 'إلى', 'الرصيد الافتتاحي', 'المفوتر', 'المسدد',
               'الصافي', 'الرصيد الختامي', 'متوسط أيام السداد']
    ws2.append(headers2)
    _style_header(ws2)
    if periodic:
        for p in periodic['periods']:
            ws2.append([p['label'], p['from'], p['to'], p['opening'], p['invoiced'],
                       p['paid'], p['net'], p['closing'],
                       p['avgSettlementDays'] if p['avgSettlementDays'] is not None else ''])
    for r in range(2, ws2.max_row + 1):
        for c in (4, 5, 6, 7, 8, 9):
            cell = ws2.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUM_FMT
    _autosize(ws2)

    # ---- الموردون
    ws3 = wb.create_sheet('الموردون')
    ws3.sheet_view.rightToLeft = True
    ws3.append(['رقم الحساب', 'الاسم', 'المشروع', 'المدة', 'المديونية القائمة', 'المتأخر'])
    _style_header(ws3)
    supplier_rows = suppliers_rows if suppliers_rows is not None else analysis.get('suppliers', [])
    for s in supplier_rows:
        ws3.append([s.get('account', ''), s.get('name', ''), s.get('project', ''),
                   s.get('term', ''), s.get('outstanding', 0), s.get('overdue', 0)])
    for r in range(2, ws3.max_row + 1):
        for c in (5, 6):
            cell = ws3.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUM_FMT
    _autosize(ws3)

    if analysis.get('contractors') is not None:
        _contractors_sheet(wb, analysis['contractors'])

    if priorities is not None:
        _priorities_sheet(wb, priorities)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
