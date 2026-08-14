# -*- coding: utf-8 -*-
"""مساعد قراءة الملفات — مزود ذكاء اصطناعي اختياري متوافق مع OpenAI.

الافتراضي Ollama محلياً بنموذج صغير، وكل شيء قابل للتعديل من الإعدادات.
الإعدادات تُحفظ كملف JSON في DATA_DIR (وليس قاعدة البيانات) حتى تنجو من
استعادة النسخ ولا تُشحن ضمن تصدير البيانات. التطبيق يعمل كاملاً بدون أي
مزود مفعّل — كل شيء هنا اختياري.
"""
import datetime as dt
import hashlib
import json
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional

import httpx

MAX_CHARS = 6000       # سقف النص المرسل في الطلب الواحد — استهلاك رموز خفيف
TIMEOUT_S = 60.0

# Neutral defaults — the user supplies their CLOUD provider's details from Settings
# (any OpenAI-compatible service). Nothing is prefilled: a wrong default endpoint
# is worse than an empty one, and the app must never assume a specific operator.
DEFAULTS = {
    'enabled': False,
    'provider': '',
    'baseUrl': '',
    'apiKey': '',
    'model': '',
    'maxTokens': 2000,
}

ARABIC_CLAUSE = (
    'أجب حصراً بالعربية الفصحى الرسمية المناسبة لمدير مالي — بلا أي كلمة إنجليزية '
    'أو نقل حرفي (transliteration) وبلا عناوين markdown. الأرقام تبقى بصيغتها الرقمية '
    'كما هي (خانات/أرقام)، لا تتحول إلى كلمات. لا تكتب أي مقدمة أو تفكير قبل الإجابة '
    '— أجب مباشرة.'
)

SYSTEM_PROMPT = (
    'أنت أداة استخراج بيانات محاسبية. يصلك نص خام من كشف حساب أو ملف مالي. '
    'استخرج سطور القيود وأعد JSON صارماً فقط بلا أي تعليق أو شرح، بالشكل: '
    '{"account": "رقم الحساب إن وجد", "name": "اسم الطرف إن وجد", '
    '"rows": [{"date": "YYYY-MM-DD", "debit": 0, "credit": 0, "description": ""}]}. '
    'التواريخ بصيغة YYYY-MM-DD حصراً، والمبالغ أرقام لا نصوص، '
    'واترك account/name غائبين إن لم تجدهما. لا تكتب أي شيء خارج الـJSON. '
    'أي نص وصفي داخل القيم (مثل description) يجب أن يكون بالعربية الفصحى، بلا '
    'كلمات إنجليزية أو markdown. ' + ARABIC_CLAUSE + '\n\n'
    # مثال واحد يرفع دقة النماذج الصغيرة كثيراً — بيانات وهمية لا تخص أي عميل حقيقي.
    'مثال — النص المُدخل:\n'
    '"مؤسسة النموذج للتجارة الرقم 1234567 رصيد افتتاحي 500.00\n'
    'فاتورة رقم 10 بتاريخ 2026-01-05 دائن 1000.00\n'
    'دفعة بتاريخ 2026-01-10 مدين 400.00"\n'
    'الناتج المطلوب لهذا المثال بالضبط:\n'
    '{"account": "1234567", "name": "مؤسسة النموذج للتجارة", '
    '"rows": [{"date": "2026-01-01", "debit": 0, "credit": 500.00, "description": "رصيد افتتاحي"}, '
    '{"date": "2026-01-05", "debit": 0, "credit": 1000.00, "description": "فاتورة رقم 10"}, '
    '{"date": "2026-01-10", "debit": 400.00, "credit": 0, "description": "دفعة"}]}'
)


_THINK_BLOCK_RE = re.compile(r'<think>.*?</think>', re.S | re.I)
# سطر يكاد يخلو من حروف عربية ويحوي حروفاً لاتينية — مقدمة إنجليزية مسرّبة من نموذج
# التفكير (reasoning) يجب حذفها قبل وصول النص للواجهة.
_ARABIC_CHAR_RE = re.compile(r'[؀-ۿ]')
_LATIN_LETTER_RE = re.compile(r'[A-Za-z]')


def _strip_leak(content: str, json_mode: bool) -> str:
    """يزيل أي تفكير/مقدمة إنجليزية مسرّبة من نموذج تفكير (reasoning model) قبل
    الاستخدام — بلا افتراض أن المزود يلتزم بالتعليمات دائماً.

    - يحذف أي كتلة <think>...</think> ظاهرة إن وُجدت.
    - في json_mode: يقتصّ النص إلى أول '{' حتى آخر '}' — أي مقدمة أو تذييل خارج
      الكائن يُطرح، لأن استدعاءات JSON لا يجوز أن تحمل نثراً حولها أصلاً.
    - في النص الحر: تُحذف الأسطر الأولى الخالية من أي حرف عربي وتحوي حروفاً لاتينية
      (مقدمة إنجليزية) حتى يصل السطر الأول الذي يحوي عربية فعلية.
    """
    text = (content or '')
    text = _THINK_BLOCK_RE.sub('', text).strip()

    if json_mode:
        start = text.find('{')
        end = text.rfind('}')
        if start != -1 and end != -1 and end > start:
            return text[start:end + 1]
        return text

    lines = text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line and not _ARABIC_CHAR_RE.search(line) and _LATIN_LETTER_RE.search(line):
            i += 1
            continue
        break
    if i > 0 and i < len(lines):
        text = '\n'.join(lines[i:]).strip()
    return text


class AiError(Exception):
    """خطأ من مزود الذكاء الاصطناعي — الرسالة عربية وصالحة للعرض مباشرة."""


# ---------------------------------------------------------------- settings

def _settings_path() -> Path:
    # import at call time so test reloads of app.core.config are honoured
    from app.core.config import settings
    return settings.DATA_DIR / 'ai-settings.json'


def load_settings() -> Dict:
    """الإعدادات المحفوظة مدموجة فوق الافتراضات — دائماً بشكل كامل."""
    out = dict(DEFAULTS)
    p = _settings_path()
    if p.exists():
        try:
            data = json.loads(p.read_text(encoding='utf-8'))
            for k in DEFAULTS:
                if k in data and data[k] is not None:
                    out[k] = data[k]
        except (ValueError, OSError):
            pass  # ملف تالف — نعود للافتراضات بدل الانهيار
    return out


def save_settings(partial: Dict) -> Dict:
    """حفظ جزئي — المفاتيح غير المذكورة تبقى كما هي."""
    cur = load_settings()
    for k in DEFAULTS:
        if k in partial and partial[k] is not None:
            cur[k] = partial[k]
    cur['maxTokens'] = int(cur['maxTokens'])
    cur['enabled'] = bool(cur['enabled'])
    p = _settings_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(cur, ensure_ascii=False, indent=2), encoding='utf-8')
    return cur


# ---------------------------------------------------------------- chat

def chat(messages: List[Dict], json_mode: bool = True) -> str:
    """POST {baseUrl}/chat/completions — درجة حرارة صفر، مهلة 60 ثانية.

    response_format=json_object عند json_mode (يدعمه Ollama)؛ إن رفضه المزود
    بـ400 نعيد المحاولة مرة واحدة بدونه.
    """
    s = load_settings()
    url = s['baseUrl'].rstrip('/') + '/chat/completions'
    headers = {'Content-Type': 'application/json'}
    if s['apiKey']:
        headers['Authorization'] = 'Bearer ' + s['apiKey']
    body = {
        'model': s['model'],
        'messages': messages,
        'temperature': 0,
        'max_tokens': int(s['maxTokens']),
    }
    if json_mode:
        body['response_format'] = {'type': 'json_object'}

    try:
        r = httpx.post(url, json=body, headers=headers, timeout=TIMEOUT_S)
        if r.status_code == 400 and json_mode:
            # مزودات ترفض response_format — محاولة ثانية بدونها
            body.pop('response_format', None)
            r = httpx.post(url, json=body, headers=headers, timeout=TIMEOUT_S)
        if r.status_code >= 400:
            # المزود عادة يشرح السبب (رصيد منتهٍ، نموذج خاطئ…) — إخفاؤه خلف رقم HTTP
            # ترك المستخدم يخمّن؛ نمرر رسالته كما هي.
            detail = ''
            try:
                body = r.json()
                detail = (body.get('error') or {}).get('message') or body.get('message') or ''
            except Exception:
                pass
            raise AiError('رفض المزود الطلب (HTTP {}){} — تأكد من العنوان والنموذج '
                          'والرصيد لدى المزود'.format(
                              r.status_code, ': ' + detail[:200] if detail else ''))
        data = r.json()
        content = data['choices'][0]['message']['content']
        return _strip_leak(content, json_mode)
    except AiError:
        raise
    except httpx.TimeoutException:
        raise AiError('انتهت مهلة الاتصال بالمزود — تأكد أن الخدمة تعمل')
    except httpx.HTTPError:
        raise AiError('تعذّر الاتصال بمزود الذكاء الاصطناعي على {} — '
                      'تأكد أن الخدمة تعمل وأن العنوان صحيح'.format(s['baseUrl']))
    except (KeyError, IndexError, ValueError):
        raise AiError('ردّ المزود بصيغة غير متوقعة — تأكد أنه متوافق مع OpenAI')


def test_connection() -> Dict:
    """اختبار سريع بأصغر طلب ممكن — لا يرفع استثناءً أبداً."""
    s = load_settings()
    try:
        chat([{'role': 'user', 'content': 'أجب بكلمة واحدة: جاهز'}], json_mode=False)
        return {'ok': True,
                'message': 'الاتصال ناجح — {} ({})'.format(s['provider'], s['model']),
                'model': s['model']}
    except AiError as e:
        return {'ok': False, 'message': str(e), 'model': s['model']}


# ---------------------------------------------------------------- extraction

def _read_pdf_pages(path: Path) -> List[str]:
    import fitz  # PyMuPDF
    doc = fitz.open(str(path))
    try:
        return [page.get_text() for page in doc]
    finally:
        doc.close()


def _read_xlsx(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    lines = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    lines.append(' | '.join(cells))
    finally:
        wb.close()
    return '\n'.join(lines)


def _clean_for_model(text: str) -> str:
    """يجهّز النص العربي قبل إرساله لأي نموذج — الخطوة التي كانت غائبة.

    هذه الكشوف تُخزّن العربية بأشكال العرض (presentation forms، مثل ﺿﻣﺎن بدل ضمان)
    — قياس فعلي على عيّنة حقيقية وجد ٢٢١ حرفاً مشوّهاً في صفحة واحدة. كل قارئ آخر في
    التطبيق (pdf_statement.py) يُطبّع النص بـNFKC قبل أي مطابقة؛ مسار الذكاء الاصطناعي
    وحده كان يرسل النص الخام — فيتلقى النموذج نصاً لا يفهمه أي نموذج مهما كان قوياً.

    NFKC يحوّل بعض هذه الأشكال إلى الياء الفارسية (ی) لا العربية (ي) — تفصيل موثّق في
    pdf_statement.py؛ غير مهم لنموذج لغوي لكنه يُوحَّد هنا لثبات المخرجات.

    كما تُحذف علامات الكشف التقنية (CompanyCode=...TrxNo=...) والأسطر الفارغة
    المتكررة — ضجيج يستهلك من حد الرموز بلا أي فائدة للنموذج.
    """
    text = unicodedata.normalize('NFKC', text)
    text = text.replace('ی', 'ي').replace('ك', 'ك')
    text = re.sub(r'CompanyCode=\S+', '', text)
    text = re.sub(r'\n[ \t]*\n+', '\n', text)
    return text.strip()


def extract_text_segments(path: Path) -> List[str]:
    """نص خام كقطع (صفحات للـPDF، نص واحد لغيره) — منظَّف ومُطبَّع لعرض عربي سليم."""
    suffix = path.suffix.lower()
    if suffix == '.pdf':
        segs = [p for p in _read_pdf_pages(path) if p.strip()]
    elif suffix in ('.xlsx', '.xlsm'):
        segs = [_read_xlsx(path)]
    else:
        segs = [path.read_text(encoding='utf-8', errors='replace')]
    return [_clean_for_model(s) for s in segs if s.strip()]


def chunk_segments(segments: List[str], limit: int = MAX_CHARS) -> List[str]:
    """تجميع القطع في دفعات لا تتجاوز الحد — قطعة أطول من الحد تُشطر."""
    pieces: List[str] = []
    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        while len(seg) > limit:
            pieces.append(seg[:limit])
            seg = seg[limit:]
        if seg:
            pieces.append(seg)
    chunks: List[str] = []
    cur = ''
    for piece in pieces:
        if cur and len(cur) + 1 + len(piece) > limit:
            chunks.append(cur)
            cur = piece
        else:
            cur = (cur + '\n' + piece) if cur else piece
    if cur:
        chunks.append(cur)
    return chunks


def _parse_json_reply(raw: str) -> Dict:
    """تحليل ردّ النموذج — نتسامح مع كتلة كود مسوّرة أو نص زائد حول الـJSON."""
    text = raw.strip()
    fenced = re.search(r'```(?:json)?\s*(.*?)```', text, re.S)
    if fenced:
        text = fenced.group(1).strip()
    try:
        return json.loads(text)
    except ValueError:
        m = re.search(r'\{.*\}', text, re.S)
        if m:
            try:
                return json.loads(m.group(0))
            except ValueError:
                pass
    raise AiError('لم يُرجع النموذج JSON صالحاً — جرّب نموذجاً آخر أو أعد المحاولة')


def _valid_rows(data: Dict) -> List[Dict]:
    rows = []
    for r in data.get('rows') or []:
        if not isinstance(r, dict):
            continue
        try:
            rows.append({
                'date': str(r.get('date') or ''),
                'debit': float(r.get('debit') or 0),
                'credit': float(r.get('credit') or 0),
                'description': str(r.get('description') or ''),
            })
        except (TypeError, ValueError):
            continue
    return rows


def extract_rows(path: Path) -> Dict:
    """استخراج سطور القيود من ملف بأي صيغة عبر النموذج — تحليل فقط، لا كتابة.

    النص يُقصّ إلى دفعات ≤ 6000 حرف، وكل دفعة تُرسل بنفس التعليمات
    وتُدمج النتائج. لا يلمس قاعدة البيانات إطلاقاً.
    """
    segments = extract_text_segments(path)
    chunks = chunk_segments(segments)
    if not chunks:
        raise AiError('الملف فارغ أو لا يحتوي نصاً يمكن قراءته')

    all_rows: List[Dict] = []
    account: Optional[str] = None
    name: Optional[str] = None
    for chunk in chunks:
        reply = chat([
            {'role': 'system', 'content': SYSTEM_PROMPT},
            {'role': 'user', 'content': chunk},
        ], json_mode=True)
        data = _parse_json_reply(reply)
        all_rows.extend(_valid_rows(data))
        if account is None and data.get('account'):
            account = str(data['account'])
        if name is None and data.get('name'):
            name = str(data['name'])

    return {
        'rows': all_rows,
        'account': account,
        'name': name,
        'chunks': len(chunks),
        'totalChars': sum(len(c) for c in chunks),
        'tokensNote': 'أُرسل {} دفعة بحد {} حرف لكل دفعة للحفاظ على استهلاك '
                      'خفيف للرموز'.format(len(chunks), MAX_CHARS),
    }


# ============================================================ learned-layout cache
#
# الفكرة: الذي يتكرر بين كشوف الحساب ليس البيانات (تتغيّر كل مرة) بل شكل الملف
# (نادراً ما يتغيّر — نفس نظام محاسبة يُصدّر نفس القالب لكل الحسابات). فبدل تخزين
# ما استُخرج (بيانات)، نخزّن كيف استُخرج (قاعدة حتمية: تعابير نمطية + ترتيب حقول).
# كشف حساب جديد بنفس القالب لا يحتاج نموذج ذكاء اصطناعي إطلاقاً — استخراج حتمي محلي.
#
# ---------------------------------------------------------------- fingerprint
#
# طريقة البصمة (compute_fingerprint) — بصراحة تامة عن حدودها:
#
# لكل ملف PDF: نأخذ أول صفحتين، نطبّعهما (NFKC + تصحيح الياء الفارسية، كما في
# _clean_for_model)، ثم نبني البصمة من: (أ) مجموعة عناوين الأعمدة/التسميات الثابتة
# الموجودة من مفردات معروفة (التاريخ/الوصف/مدين/دائن/الحساب/الرقم…)، (ب) وجود علامة
# "CompanyCode=" التقنية التي يضعها نظام تصدير معيّن. عمداً **لا** تدخل بصمة PDF وجود
# سطر "رصيد افتتاحي" ضمن الحساب — رغم أن هذا مذكور كمثال في التصميم الأصلي، القياس
# على عيّنات حقيقية (design/samples) أظهر أن حسابين من *نفس* النظام قد يختلفان في
# وجود هذا السطر (يعتمد على تاريخ بداية الحساب لا على قالب التصدير) — تضمينه كان
# يعطي بصمتين مختلفتين لنفس القالب، وهو بالضبط الخطأ الذي لا نريده (false negative).
#
# لكل ملف xlsx: البصمة = أنماط أسماء الأوراق (الأرقام تُستبدل بـ# حتى "Sheet1" و
# "Sheet2" يتطابقان) + محتوى أول صف يحمل خليتين فأكثر (صف العناوين على الأرجح) من
# أول ورقة، مطبَّعاً.
#
# لغير ذلك (csv/txt): بصمة خشنة من عدد الفواصل في أول سطر — تغطية دنيا فقط؛ المسار
# الحتمي عملياً لن يفعّل لهذه الصيغة لأن لا قواعد تُتعلَّم لها بعد.
#
# المفاضلات الصادقة:
# - False positive (بصمتان متطابقتان لقالبين مختلفين فعلياً): وارد — عناوين الأعمدة
#   العربية الشائعة (التاريخ/الوصف/مدين/دائن) تتكرر بين أنظمة محاسبة مختلفة تماماً؛
#   إن حمل قالب آخر نفس التسميات الأربع بلا علامة CompanyCode فقد تتصادم بصمتاهما.
#   النتيجة عند الخطأ: محاولة استخراج حتمي تفشل معيار المعقولية (_plausible) فتتراجع
#   تلقائياً لمسار النموذج — لا خطأ صامت، فقط فرصة توفير ضائعة.
# - False negative (نفس القالب فعلياً لكن بصمتان مختلفتان): وارد إن تغيّرت تسميات
#   الأعمدة بين إصدارين من نفس النظام، أو كانت الصفحة الأولى تحوي نصاً إضافياً غير
#   معتاد يُخفي بعض التسميات. النتيجة عند الخطأ: تعلّم مكرر (صفّان بدل صف) — غير
#   خطير، فقط توفير أقل من الأمثل.
#
# القاعدة الأهم تبقى في مكان آخر (extract_rows/routes): البصمة تحدد فقط *أي قاعدة
# حتمية نجرّب*، ولا تُستخدم أبداً لتحديد ما يُكتب في قاعدة البيانات مباشرة.

_PDF_HEADER_VOCAB = ('التاريخ', 'الوصف', 'مدين', 'دائن', 'الحساب', 'الرقم', 'رصيد', 'الفرع')
_ARABIC_LETTER_RE = re.compile(r'[ء-ي]')
_DATE_DMY_RE = re.compile(r'\b(\d{2})-(\d{2})-(\d{4})\b')
_AMOUNT_RE = re.compile(r'\(?\d{1,3}(?:,\d{3})*\.\d{2}\)?')


def _normalize_for_rules(text: str) -> str:
    """نفس تطبيع NFKC + تصحيح الياء الفارسية في _clean_for_model، لكن **بدون** حذف
    علامة CompanyCode= أو ضغط الأسطر الفارغة — القواعد الحتمية تحتاجهما كعلامات
    فاصلة بين سطور القيود، بعكس ما يُرسل فعلياً للنموذج."""
    text = unicodedata.normalize('NFKC', text)
    return text.replace('ی', 'ي')


def source_kind_of(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == '.pdf':
        return 'pdf'
    if suffix in ('.xlsx', '.xlsm'):
        return 'xlsx'
    return 'other'


def compute_fingerprint(path: Path) -> str:
    """بصمة شكل الملف (لا محتواه) — انظر الشرح الطويل أعلاه لحدودها الصادقة."""
    kind = source_kind_of(path)
    if kind == 'pdf':
        pages = _read_pdf_pages(path)
        sample_raw = '\n'.join(pages[:2])
        has_marker = 'CompanyCode=' in sample_raw
        normalized = _normalize_for_rules(sample_raw)
        labels = sorted(lbl for lbl in _PDF_HEADER_VOCAB if lbl in normalized)
        raw_input = 'pdf|labels=' + ','.join(labels) + '|marker=' + str(has_marker)
    elif kind == 'xlsx':
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            sheet_names = [re.sub(r'\d+', '#', ws.title) for ws in wb.worksheets]
            header_cells: List[str] = []
            if wb.worksheets:
                for row in wb.worksheets[0].iter_rows(values_only=True):
                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if len(cells) >= 2:
                        header_cells = cells
                        break
        finally:
            wb.close()
        normalized_header = [_normalize_for_rules(c) for c in header_cells]
        raw_input = ('xlsx|sheets=' + ','.join(sheet_names) + '|header='
                    + '|'.join(normalized_header) + '|cols=' + str(len(header_cells)))
    else:
        text = path.read_text(encoding='utf-8', errors='replace')
        first_line = text.split('\n', 1)[0] if text else ''
        raw_input = 'other|delims=' + str(first_line.count(','))
    return hashlib.sha256(raw_input.encode('utf-8')).hexdigest()


# ---------------------------------------------------------------- rules: learn

def _raw_text_for_rules(path: Path) -> str:
    """نص خام (بدون تنظيف النموذج) مُطبَّع فقط — أساس تعلّم القواعد وتطبيقها."""
    kind = source_kind_of(path)
    if kind == 'pdf':
        raw = '\n'.join(_read_pdf_pages(path))
    elif kind == 'xlsx':
        raw = _read_xlsx(path)
    else:
        raw = path.read_text(encoding='utf-8', errors='replace')
    return _normalize_for_rules(raw)


def learn_rules_from_extraction(raw_text: str, result: Dict) -> Optional[Dict]:
    """يستنتج قاعدة استخراج حتمية من نص خام + نتيجة نموذج ناجحة — كودياً بالكامل،
    بلا أي استدعاء إضافي للنموذج (تفضيلاً معلناً في التصميم).

    الأنماط المُستنتَجة ترتكز على تسميات أعمدة ثابتة (التاريخ/الوصف/الحساب/الرقم)
    لا على قيم البيانات نفسها (رقم الحساب، التاريخ الفعلي…) — هذا ما يجعلها قابلة
    لإعادة الاستخدام على ملف آخر من نفس النظام ببيانات مختلفة تماماً.

    يعيد None إن لم يجد على الأقل صفاً واحداً يمكن ربطه بنمط تاريخ في النص — عندها
    لا تُحفظ قاعدة، والملف التالي من نفس القالب يمر مجدداً بالنموذج.
    """
    rows = result.get('rows') or []
    if not rows:
        return None

    row_split_regex = r'CompanyCode=\S+' if 'CompanyCode=' in raw_text else None

    # تسميات ثابتة تسبق قيمة البيانات في ترتيب الاستخراج المعكوس لملفات PDF بهذا
    # النظام — رقم الحساب يظهر مباشرة قبل سطر ": الحساب"، واسم الطرف قبل "الرقم".
    account_regex = None
    if re.search(r'\d{3,9}\s*\n\s*:\s*الحساب', raw_text):
        account_regex = r'(\d{3,9})\s*\n\s*:\s*الحساب'
    name_regex = None
    if re.search(r'[^\n]+\n(?:الرقم)', raw_text):
        name_regex = r'([^\n]+)\n(?:الرقم)'

    date_format = None
    for r in rows:
        d = str(r.get('date') or '')
        if re.match(r'^\d{2}-\d{2}-\d{4}$', d) and d in raw_text:
            date_format = 'DD-MM-YYYY'
            break
    if date_format is None and _DATE_DMY_RE.search(raw_text):
        date_format = 'DD-MM-YYYY'  # النمط الوحيد المدعوم حالياً لهذا النظام

    if date_format is None:
        return None  # لا نمط تاريخ معروف يمكن الاعتماد عليه لملف لاحق

    rules = {
        'row_split_regex': row_split_regex,
        'account_regex': account_regex,
        'name_regex': name_regex,
        'date_regex': _DATE_DMY_RE.pattern,
        'date_format': date_format,
        'amount_regex': _AMOUNT_RE.pattern,
        'amount_order': ['debit', 'credit'],
    }
    return rules


def _last_description_line(block: str) -> str:
    lines = [ln.strip() for ln in block.split('\n') if ln.strip()]
    for ln in reversed(lines):
        if _ARABIC_LETTER_RE.search(ln) and not _DATE_DMY_RE.fullmatch(ln.strip()):
            return re.sub(r'^0*\d{0,4}(?=\D)', '', ln).strip() or ln
    return ''


def _parse_amount(raw: str) -> float:
    neg = raw.startswith('(') and raw.endswith(')')
    cleaned = raw.strip('()').replace(',', '')
    try:
        val = float(cleaned)
    except ValueError:
        return 0.0
    return -val if neg else val


def apply_rules(raw_text: str, rules: Dict) -> Optional[Dict]:
    """يطبّق قاعدة مُتعلَّمة على نص خام لملف جديد — استخراج حتمي محلي، صفر رموز."""
    account = None
    if rules.get('account_regex'):
        m = re.search(rules['account_regex'], raw_text)
        if m:
            account = m.group(1)
    name = None
    if rules.get('name_regex'):
        m = re.search(rules['name_regex'], raw_text)
        if m:
            name = m.group(1).strip()

    date_re = re.compile(rules.get('date_regex') or _DATE_DMY_RE.pattern)
    amount_re = re.compile(rules.get('amount_regex') or _AMOUNT_RE.pattern)
    order = rules.get('amount_order') or ['debit', 'credit']
    date_format = rules.get('date_format') or 'DD-MM-YYYY'

    split_regex = rules.get('row_split_regex')
    blocks = re.split(split_regex, raw_text)[1:] if split_regex else [raw_text]

    rows: List[Dict] = []
    for raw_block in blocks:
        # القطعة تمتد حتى بداية القطعة التالية أو نهاية الصفحة — أي نص إجمالي/تذييل
        # بعد آخر قيد قد يتسرب لوصف آخر سطر؛ "Page" علامة تذييل ثابتة في هذا النظام.
        block = raw_block.split('Page ')[0]
        dm = date_re.search(block)
        if not dm:
            continue
        if date_format == 'DD-MM-YYYY':
            try:
                iso = dt.date(int(dm.group(3)), int(dm.group(2)), int(dm.group(1))).isoformat()
            except (ValueError, IndexError):
                continue
        else:
            continue
        amounts = amount_re.findall(block)
        if not amounts:
            continue
        values = {k: 0.0 for k in ('debit', 'credit')}
        for key, raw_amt in zip(order, amounts):
            if key in values:
                values[key] = abs(_parse_amount(raw_amt))
        rows.append({
            'date': iso,
            'debit': values['debit'],
            'credit': values['credit'],
            'description': _last_description_line(block),
        })

    if not rows:
        return None
    return {'rows': rows, 'account': account, 'name': name}


def _plausible(result: Optional[Dict]) -> bool:
    """معيار المعقولية — يقرر إن كانت محاولة الاستخراج الحتمي صالحة للعرض فوراً
    أو يجب التراجع لمسار النموذج: حساب أو اسم موجود، وصف واحد على الأقل، وكل
    التواريخ/المبالغ صالحة (وهذا مضمون فعلياً من apply_rows نفسها)."""
    if not result:
        return False
    if not result.get('rows'):
        return False
    if not result.get('account') and not result.get('name'):
        return False
    for r in result['rows']:
        try:
            dt.date.fromisoformat(r['date'])
            float(r['debit'])
            float(r['credit'])
        except (ValueError, TypeError, KeyError):
            return False
    return True


# ---------------------------------------------------------------- deterministic path

def try_deterministic_extract(db, path: Path) -> Optional[Dict]:
    """محاولة استخراج حتمي محلي — صفر رموز — قبل أي استدعاء للنموذج.

    تُستدعى من /ai/extract قبل extract_rows. تعيد None إن لم توجد قاعدة متعلَّمة
    لهذا الشكل أو لم تنجح تطبيقها بمعيار المعقولية _plausible — عندها يستمر مسار
    الاستدعاء العادي بلا أي أثر جانبي. **لا تكتب في قاعدة البيانات المالية أبداً**؛
    فقط تقرأ/تحدّث جدول learned_layouts (عدّاد الاستخدام)، ونتيجتها تمر كأي استخراج
    آخر عبر المراجعة البشرية الإلزامية في /ai/commit-extract.
    """
    from app.db import models

    fingerprint = compute_fingerprint(path)
    layout = db.query(models.LearnedLayout).filter_by(fingerprint=fingerprint).one_or_none()
    if layout is None:
        return None

    try:
        rules = json.loads(layout.rules_json or '{}')
    except ValueError:
        return None

    raw_text = _raw_text_for_rules(path)
    result = apply_rules(raw_text, rules)
    if not _plausible(result):
        return None

    layout.hit_count = (layout.hit_count or 0) + 1
    layout.last_used_at = dt.datetime.now(dt.timezone.utc)
    db.commit()

    approx_tokens_saved = max(len(raw_text) // 4, 1)
    result['chunks'] = 0
    result['source'] = 'learned'
    result['layoutHitCount'] = layout.hit_count
    result['tokensSaved'] = approx_tokens_saved
    result['tokensNote'] = 'استُخدمت قاعدة تخطيط متعلَّمة سابقاً — لم يُستدعَ أي نموذج ذكاء اصطناعي'
    return result


def learn_from_extraction(db, path: Path, sent_chars: int, result: Dict) -> None:
    """بعد نجاح استخراج بالنموذج: تستنتج قاعدة حتمية كودياً (بلا استدعاء نموذج
    إضافي) وتحفظها/تحدّثها في learned_layouts.

    إن كانت البصمة موجودة مسبقاً، هذا يعني منطقياً أن المسار الحتمي كان يجب أن
    يُفعَّل قبل الوصول للنموذج ولم يفعل — يُسجَّل هذا كتحذير لتحسين القواعد لاحقاً
    بدل تجاهله صامتاً، ويُكتفى بزيادة hitCount دون الكتابة فوق قاعدة موجودة (قد
    تكون أدق من إعادة الاستنتاج الآلي من عيّنة واحدة).
    """
    import logging
    from app.db import models

    logger = logging.getLogger(__name__)
    fingerprint = compute_fingerprint(path)
    existing = db.query(models.LearnedLayout).filter_by(fingerprint=fingerprint).one_or_none()

    if existing is not None:
        existing.hit_count = (existing.hit_count or 0) + 1
        existing.last_used_at = dt.datetime.now(dt.timezone.utc)
        db.commit()
        logger.warning(
            'learned-layout fingerprint %s already known but the deterministic path did '
            'not fire before the LLM call — check apply_rules against this file to tune '
            'the stored rules', fingerprint)
        return

    raw_text = _raw_text_for_rules(path)
    rules = learn_rules_from_extraction(raw_text, result)
    if rules is None:
        return  # لا نمط يمكن الوثوق به — لا تُحفظ قاعدة جزئية غير موثوقة

    layout = models.LearnedLayout(
        fingerprint=fingerprint,
        source_kind=source_kind_of(path),
        sample_account=(result.get('account') or '')[:32],
        sample_name=(result.get('name') or '')[:300],
        rules_json=json.dumps(rules, ensure_ascii=False),
        hit_count=0,
        learned_from_chars=sent_chars,
    )
    db.add(layout)
    db.commit()
