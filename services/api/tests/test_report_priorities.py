# -*- coding: utf-8 -*-
"""اختبارات قسم «أولويات السداد» داخل التقرير التحليلي.

- GET /api/v1/reports/priorities: نفس الترتيب الحتمي (F.build_priorities) لكن بلا
  استدعاء للمساعد الآلي أبداً — يجب أن يعمل حتى مع إعدادات AI متوقفة (الحالة
  الافتراضية في بيئة الاختبار).
- تصدير Excel الرئيسي (export.xlsx) يجب أن يحمل ورقة «أولويات السداد» بنفس
  الأرقام التي يُعيدها المسار أعلاه — لا رقم يُخترع ولا صف يُسقَط بصمت.
"""
import datetime as dt

from openpyxl import load_workbook
import io

BASE = '/api/v1/reports'


def seed_overdue_supplier(client, account='9101', name='مورد متأخر',
                          project='مشروع الأولويات', term='30 يوم'):
    r = client.post('/api/v1/suppliers', json={
        'account': account, 'name': name, 'project': project, 'term': term})
    assert r.status_code in (200, 201), r.text
    # فاتورة قديمة كفاية لتكون متأخرة بيقين بغضّ النظر عن تاريخ اليوم
    r = client.post('/api/v1/manual/invoices', json={
        'account': account, 'date': '2020-01-01', 'amount': 5000, 'number': 'PF1'})
    assert r.status_code in (200, 201), r.text
    return account


def test_priorities_endpoint_works_without_ai(api_client):
    """المسار الحتمي لا يعتمد على تفعيل AI — يعمل ويُرجع عناصر ومبالغ صحيحة."""
    account = seed_overdue_supplier(api_client)

    r = api_client.get(f'{BASE}/priorities')
    assert r.status_code == 200, r.text
    d = r.json()
    assert 'items' in d and 'narrative' not in d  # لا نص آلي في هذا المسار إطلاقاً
    assert any(it['key'] == account for it in d['items'])
    item = next(it for it in d['items'] if it['key'] == account)
    assert item['partyKind'] == 'supplier'
    assert item['amount'] == 5000.0
    assert item['score'] > 0


def test_priorities_endpoint_budget_filter_shape(api_client):
    """تمرير ميزانية لا يغيّر شكل الاستجابة الأساسي — budget_info يظهر فقط حينها."""
    seed_overdue_supplier(api_client)

    r_no_budget = api_client.get(f'{BASE}/priorities')
    assert r_no_budget.json().get('budget') is None

    r_budget = api_client.get(f'{BASE}/priorities', params={'budget': '1000'})
    assert r_budget.status_code == 200, r_budget.text
    d = r_budget.json()
    assert d['budget'] is not None
    assert d['budget']['budget'] == 1000.0


def test_priorities_endpoint_empty_when_nothing_overdue(api_client):
    """لا موردين متأخرين ولا مقاولين بأرصدة سالبة -> قائمة فارغة، لا خطأ."""
    r = api_client.get(f'{BASE}/priorities')
    assert r.status_code == 200, r.text
    assert r.json()['items'] == []


def test_export_xlsx_includes_priorities_sheet_matching_endpoint(api_client):
    """ورقة «أولويات السداد» في التصدير يجب أن تحمل نفس الاسم والمبلغ والسبب
    الظاهرة في GET /api/v1/reports/priorities — نفس الأرقام حرفياً، لا تقريب مختلف.
    """
    account = seed_overdue_supplier(api_client)

    ref = api_client.get(f'{BASE}/priorities').json()
    ref_item = next(it for it in ref['items'] if it['key'] == account)

    r = api_client.get(f'{BASE}/export.xlsx')
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content))
    assert 'أولويات السداد' in wb.sheetnames
    ws = wb['أولويات السداد']

    header = [c.value for c in ws[1]]
    assert header == ['#', 'الاسم', 'النوع', 'المبلغ (ر.س)', 'السبب']

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    matching = [row for row in rows if row[1] == ref_item['name']]
    assert matching, 'صف المورد المتأخر غائب عن ورقة أولويات السداد المصدَّرة'
    row = matching[0]
    assert row[2] == 'مورد'
    assert float(row[3]) == ref_item['amount']
    assert row[4] == ref_item['reason']
