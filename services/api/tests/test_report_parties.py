# -*- coding: utf-8 -*-
"""اختبارات نطاق أطراف التقرير — suppliers / contractors / both.

Every figure asserted here is hand-computed from the seeded fixtures below, so a change
in the contractor→report mapping (Σcredit = invoiced, payment debits = paid, the rest =
deductions) fails loudly instead of silently re-labelling money.
"""
from decimal import Decimal

BASE = '/api/v1/reports'

#: keys the report has always had — a default (parties omitted) call must still be
#: exactly this shape, no key removed, so old frontends keep rendering.
LEGACY_TOP_KEYS = {'meta', 'summary', 'ageing', 'schedule', 'suppliers', 'notes'}
LEGACY_SUMMARY_KEYS = {'total_invoiced', 'total_paid', 'outstanding', 'overdue',
                       'due_within_7', 'supplier_count',
                       'credit_balances', 'net_outstanding'}


# ---------------------------------------------------------------- seeding

def seed_supplier(client, account='9001', term='30 يوم'):
    r = client.post('/api/v1/suppliers', json={
        'account': account, 'name': 'مورد التجربة', 'project': 'السدن', 'term': term})
    assert r.status_code in (200, 201), r.text
    r = client.post('/api/v1/manual/invoices', json={
        'account': account, 'date': '2026-01-10', 'amount': 1000, 'number': 'F1'})
    assert r.status_code in (200, 201), r.text
    r = client.post('/api/v1/manual/payments', json={
        'account': account, 'date': '2026-02-10', 'amount': 400})
    assert r.status_code in (200, 201), r.text


def seed_contractors(client):
    """مقاولان: واحد ندين له و واحد يدين لنا — both signs must be exercised."""
    client.post('/api/v1/contractors', json={'code': 'C1', 'name': 'مقاول أول'})
    client.post('/api/v1/contractors', json={'code': 'C2', 'name': 'مقاول ثانٍ'})

    # C1: credit 5000 (مستخلص) ; debit 1200 payment + 300 خصم  →  balance = -3500
    client.post('/api/v1/contractors/C1/entries', json={
        'date': '2026-01-05', 'credit': 5000, 'description': 'مستخلص رقم1',
        'project': 'السدن'})
    client.post('/api/v1/contractors/C1/entries', json={
        'date': '2026-02-05', 'debit': 1200, 'description': 'دفعة نقدية',
        'project': 'السدن'})
    client.post('/api/v1/contractors/C1/entries', json={
        'date': '2026-02-06', 'debit': 300, 'description': 'خصم مواد',
        'project': 'السدن'})

    # C2: credit 1000 ; debit 1500 payment  →  balance = +500 (he owes us)
    client.post('/api/v1/contractors/C2/entries', json={
        'date': '2026-01-07', 'credit': 1000, 'description': 'مستخلص رقم1'})
    client.post('/api/v1/contractors/C2/entries', json={
        'date': '2026-03-07', 'debit': 1500, 'description': 'دفعة نقدية'})


# ---------------------------------------------------------------- regression guard

def test_default_parties_keeps_the_legacy_payload_shape(api_client):
    seed_supplier(api_client)
    seed_contractors(api_client)

    a = api_client.get(f'{BASE}/analysis').json()
    b = api_client.get(f'{BASE}/analysis?parties=suppliers').json()

    assert set(a) == LEGACY_TOP_KEYS
    assert 'contractors' not in a
    assert set(a['summary']) == LEGACY_SUMMARY_KEYS
    assert a['meta']['scope_label'] == 'كل الموردين'
    assert a['meta']['title'] == 'تقرير تحليل مديونية الموردين'
    # contractors present in the database must not leak into a supplier-only report
    assert a['summary']['outstanding'] == 600.0
    assert a == b


def test_supplier_rows_and_schedule_name_their_party(api_client):
    seed_supplier(api_client)
    p = api_client.get(f'{BASE}/analysis').json()
    row = p['suppliers'][0]
    assert row['partyKind'] == 'supplier'
    assert row['account'] == '9001' and row['name'] == 'مورد التجربة'
    for day in p['schedule']:
        for it in day['items']:
            assert it['partyKind'] == 'supplier'
            assert it['name'] and it['account']


# ---------------------------------------------------------------- contractors

def test_contractors_only_scope_maps_the_ledger_into_report_words(api_client):
    seed_supplier(api_client)
    seed_contractors(api_client)
    p = api_client.get(f'{BASE}/analysis?parties=contractors').json()

    t = p['contractors']['totals']
    assert t['count'] == 2
    assert t['invoiced'] == 6000.0                      # 5000 + 1000 credits
    assert t['paid'] == 2700.0                          # 1200 + 1500 payment debits
    assert t['deductions'] == 300.0                     # the خصم debit, never "paid"
    assert t['balance'] == -3000.0                      # (1500-5000) + (1500-1000)
    assert t['outstanding'] == 3500.0                   # only the negative one counts

    # suppliers are out of scope: their numbers must be absent from the summary
    assert p['summary']['supplier_count'] == 0
    assert p['summary']['total_invoiced'] == 6000.0
    assert p['summary']['outstanding'] == 3500.0
    assert p['meta']['scope_label'] == 'كل المقاولين'
    assert p['meta']['parties'] == 'contractors'
    assert p['meta']['includes'] == dict(suppliers=False, contractors=True)
    # no faked ageing for contractors
    assert p['meta']['ageing_covers'] == 'suppliers'
    assert all(row['amount'] == 0 for row in p['ageing'])
    assert p['schedule'] == []

    rows = {r['code']: r for r in p['contractors']['rows']}
    assert rows['C1']['partyKind'] == 'contractor'
    assert rows['C1']['projects'] == ['السدن']
    assert rows['C1']['lastPayment']['amount'] == 1200.0
    assert rows['C2']['outstanding'] == 0.0 and rows['C2']['balance'] == 500.0
    # most-negative first
    assert [r['code'] for r in p['contractors']['rows']] == ['C1', 'C2']


def test_both_mode_is_exactly_supplier_plus_contractor_totals(api_client):
    seed_supplier(api_client)
    seed_contractors(api_client)

    s = api_client.get(f'{BASE}/analysis').json()['summary']
    c = api_client.get(f'{BASE}/analysis?parties=contractors').json()['contractors']['totals']
    both = api_client.get(f'{BASE}/analysis?parties=both').json()

    d = lambda v: Decimal(str(v))
    assert d(both['summary']['total_invoiced']) == d(s['total_invoiced']) + d(c['invoiced'])
    assert d(both['summary']['total_paid']) == d(s['total_paid']) + d(c['paid'])
    assert d(both['summary']['outstanding']) == d(s['outstanding']) + d(c['outstanding'])
    # ageing/overdue stay supplier-only in both mode
    assert both['summary']['overdue'] == s['overdue']
    assert both['summary']['supplier_count'] == s['supplier_count']
    assert both['summary']['contractor_count'] == 2
    assert both['meta']['scope_label'] == 'الموردون والمقاولون'
    assert both['meta']['includes'] == dict(suppliers=True, contractors=True)


def test_bad_parties_value_is_rejected(api_client):
    assert api_client.get(f'{BASE}/analysis?parties=nobody').status_code == 422


# ---------------------------------------------------------------- single contractor

def test_single_contractor_scope(api_client):
    seed_supplier(api_client)
    seed_contractors(api_client)
    p = api_client.get(f'{BASE}/analysis?contractor=C1').json()

    assert p['meta']['scope'] == 'contractor'
    assert p['meta']['scope_label'] == 'المقاول: مقاول أول'
    assert p['contractors']['totals']['count'] == 1
    assert p['contractors']['totals']['outstanding'] == 3500.0
    by_project = p['contractorDetail']['byProject']
    assert [b['project'] for b in by_project] == ['السدن']
    assert by_project[0]['entryCount'] == 3
    assert all(e['partyKind'] == 'contractor' and e['code'] == 'C1'
               for e in by_project[0]['entries'])


def test_unknown_contractor_scope_is_404(api_client):
    assert api_client.get(f'{BASE}/analysis?contractor=NOPE').status_code == 404
    assert api_client.get(f'{BASE}/export.xlsx?contractor=NOPE').status_code == 404


# ---------------------------------------------------------------- scopes + export

def test_scopes_lists_contractors(api_client):
    seed_supplier(api_client)
    seed_contractors(api_client)
    s = api_client.get(f'{BASE}/scopes').json()
    assert s['parties'] == ['suppliers', 'contractors', 'both']
    codes = {c['code']: c for c in s['contractors']}
    assert set(codes) == {'C1', 'C2'}
    assert codes['C1']['hasData'] is True and codes['C1']['name'] == 'مقاول أول'
    assert [x['account'] for x in s['suppliers']] == ['9001']


def test_export_sheets_follow_the_party_scope(api_client):
    import io
    from openpyxl import load_workbook

    seed_supplier(api_client)
    seed_contractors(api_client)

    # ورقة «أولويات السداد» تُدرَج في التصدير العام فقط. قائمتها على مستوى الشركة
    # بالكامل، فوجودها داخل تصدير مقاول بعينه يجعل الوثيقة تكذب بنطاقها.
    r = api_client.get(f'{BASE}/export.xlsx')
    assert r.status_code == 200
    assert load_workbook(io.BytesIO(r.content)).sheetnames == [
        'الملخص', 'الفترات', 'الموردون', 'أولويات السداد']

    r = api_client.get(f'{BASE}/export.xlsx?parties=both')
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ['الملخص', 'الفترات', 'الموردون', 'المقاولون', 'أولويات السداد']
    assert wb['المقاولون'].max_row == 4          # header + 2 contractors + totals

    r = api_client.get(f'{BASE}/export.xlsx?contractor=C1')
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ['المقاولون']   # نطاق مقاول واحد — بلا أولويات
    assert 'contractor-C1' in r.headers['content-disposition']
