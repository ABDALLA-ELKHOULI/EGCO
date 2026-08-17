# -*- coding: utf-8 -*-
"""بحث وتصفية عربيان مُطبَّعان — الموردون، المقاولون، ومشاريع الأطراف.

يثبت السيناريو الحقيقي المُبلَّغ عنه: البحث بالإملاء «الصحيح» (تاء مربوطة، همزة،
ياء عادية) لا يجد مورداً أُدخل اسمه بإملاء آخر (هاء، ألف بلا همزة، ياء فارسية)
رغم أنه نفس الكيان بالضبط — وأن تصفية المشروع تعاني نفس العلة فتُخفي طرفاً
مسجَّلاً تحت صيغة إملائية أخرى لنفس المشروع.

يبني بياناته عبر مسارات الـ API نفسها، بنفس نمط test_contractors_filters.py.
"""
import importlib

import pytest

from app.utils.arabic import contains_ar, normalize_ar, same_ar


# ---------------------------------------------------------------- unit: normalize_ar

def test_normalize_hamza_forms_unify_to_alef():
    assert normalize_ar('أحمد') == normalize_ar('احمد') == normalize_ar('إحمد') == normalize_ar('آحمد')


def test_normalize_taa_marbuta_and_heh():
    assert normalize_ar('الهضبة') == normalize_ar('الهضبه')
    assert normalize_ar('شركة قنبر') == normalize_ar('شركه قنبر')


def test_normalize_alef_maqsura_and_farsi_yeh():
    assert normalize_ar('بيت الآباء') == normalize_ar('بيت الاباء')
    assert normalize_ar('على') == normalize_ar('علي')  # ى → ي


def test_normalize_tashkeel_and_tatweel_stripped():
    assert normalize_ar('مُحَمَّد') == normalize_ar('محمد')
    assert normalize_ar('كـبير') == normalize_ar('كبير')


def test_normalize_empty_and_none_safe():
    assert normalize_ar('') == ''
    assert normalize_ar(None) == ''


def test_contains_ar_matches_across_spelling_variants():
    assert contains_ar('شركة الهضبه الصناعيه الروشن', 'الهضبة')
    assert not contains_ar('شركة الهضبه الصناعيه الروشن', 'الغربية')


def test_contains_ar_empty_needle_matches_everything():
    assert contains_ar('أي اسم', '')
    assert contains_ar('أي اسم', '   ')


def test_same_ar():
    assert same_ar('المدينة', 'المدينه')
    assert not same_ar('المدينة', 'الرياض')


# ---------------------------------------------------------------- suppliers API

@pytest.fixture()
def env(tmp_path, monkeypatch):
    monkeypatch.setenv('EGCO_DATA_DIR', str(tmp_path / 'data'))

    import app.core.config as config_mod
    importlib.reload(config_mod)
    import app.db.session as session_mod
    importlib.reload(session_mod)
    import app.db.models as models_mod
    importlib.reload(models_mod)
    import app.services.payables_service as payables_service_mod
    importlib.reload(payables_service_mod)
    import app.services.party_projects as party_projects_mod
    importlib.reload(party_projects_mod)
    import app.services.contractors_service as contractors_service_mod
    importlib.reload(contractors_service_mod)
    import app.api.routes.suppliers as suppliers_route
    importlib.reload(suppliers_route)
    import app.api.routes.contractors as contractors_route
    importlib.reload(contractors_route)

    session_mod.init_db()

    class Env:
        pass

    e = Env()
    e.session = session_mod
    e.suppliers_route = suppliers_route
    e.contractors_route = contractors_route
    return e


@pytest.fixture()
def client(env):
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    app = FastAPI()
    app.include_router(env.suppliers_route.router, prefix='/suppliers')
    app.include_router(env.contractors_route.router, prefix='/contractors')
    with TestClient(app) as c:
        yield c


def test_supplier_search_finds_correct_spelling_when_stored_differently(client):
    # البيانات كما صدَّرها نظام الحسابات القديم — تاء مربوطة صارت هاء.
    r = client.post('/suppliers', json=dict(
        account='2111724', name='شركة الهضبه الصناعيه الروشن', project='الرشين'))
    assert r.status_code == 201

    # المستخدم يبحث بالإملاء الصحيح إملائياً (تاء مربوطة) — يجب أن يجده رغم
    # اختلاف الحرف الأخير عن المخزَّن بالضبط.
    found = client.get('/suppliers?q=الهضبة').json()
    assert found['count'] == 1
    assert found['rows'][0]['account'] == '2111724'


def test_supplier_search_progressive_typing_does_not_vanish(client):
    """يعيد سيناريو المستخدم بالحرف: كتابة الاسم تدريجياً حتى الحرف المختلف
    إملائياً لا يجب أن تُسقط النتيجة إلى صفر."""
    client.post('/suppliers', json=dict(
        account='2111724', name='شركة الهضبه الصناعيه الروشن', project='الرشين'))
    for prefix in ('ا', 'اله', 'الهض', 'الهضب', 'الهضبة', 'الهضبة الصناعية'):
        r = client.get(f'/suppliers?q={prefix}').json()
        assert r['count'] == 1, f'اختفت النتيجة عند "{prefix}"'


def test_supplier_search_by_hamza_and_yeh_variants(client):
    client.post('/suppliers', json=dict(
        account='2110963', name='بيت الاباء مشروع روشن(مواد سباكه)', project='الرشين'))
    r = client.get('/suppliers?q=بيت الآباء').json()
    assert r['count'] == 1


def test_supplier_project_filter_groups_orthographic_variants(client):
    # مورد أول على «المدينة»، والثاني أُدخل مشروعه «المدينه» — نفس المشروع بالضبط
    # بإملاء مختلف حرفاً واحداً، كما يحدث فعلاً عند إدخال يدوي غير متسق.
    client.post('/suppliers', json=dict(account='111', name='مورد أول', project='المدينة'))
    client.post('/suppliers', json=dict(account='222', name='مورد ثاني', project='المدينه'))

    listing = client.get('/suppliers').json()
    # لائحة المشاريع تُظهر صيغة واحدة لا اثنتين لنفس الهوية.
    assert len([p for p in listing['projects'] if same_ar_helper(p, 'المدينة')]) == 1

    by_taa = client.get('/suppliers?project=المدينة').json()
    by_heh = client.get('/suppliers?project=المدينه').json()
    assert by_taa['count'] == 2
    assert by_heh['count'] == 2


def same_ar_helper(a, b):
    return same_ar(a, b)


def test_contractor_search_finds_correct_spelling(client):
    client.post('/contractors', json=dict(code='C1', name='شركة قنبر للخرسانة الجاهزه'))
    r = client.get('/contractors?q=شركه قنبر').json()
    assert r['count'] == 1
    assert r['rows'][0]['code'] == 'C1'


def test_contractor_project_filter_groups_orthographic_variants(client):
    client.post('/contractors', json=dict(code='C1', name='مقاول أول', projects=['المدينة']))
    client.post('/contractors', json=dict(code='C2', name='مقاول ثاني', projects=['المدينه']))
    r = client.get('/contractors?project=المدينة').json()
    codes = {row['code'] for row in r['rows']}
    assert codes == {'C1', 'C2'}


def test_supplier_export_reflects_filtered_set_only(client):
    client.post('/suppliers', json=dict(account='111', name='مورد أول', project='الرشين'))
    client.post('/suppliers', json=dict(account='222', name='مورد ثاني', project='القصر'))

    r = client.get('/suppliers/export.xlsx?project=الرشين')
    assert r.status_code == 200
    assert r.headers['content-type'].startswith(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(r.content))
    ws = wb.active
    accounts = [row[0].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1)]
    assert accounts == ['111']


def test_contractor_export_reflects_filtered_set_only(client):
    client.post('/contractors', json=dict(code='C1', name='مقاول أول', projects=['الرشين']))
    client.post('/contractors', json=dict(code='C2', name='مقاول ثاني', projects=['القصر']))

    r = client.get('/contractors/export.xlsx?project=الرشين')
    assert r.status_code == 200
    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(r.content))
    ws = wb.active
    codes = [row[0].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1)]
    assert codes == ['C1']
